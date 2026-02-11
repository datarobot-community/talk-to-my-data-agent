# Copyright 2025 DataRobot, Inc.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Router for chat endpoints."""

from __future__ import annotations

import io
import json
import os
import tempfile
from typing import Any, List, Union, cast

import pandas as pd
import polars.dataframe.frame
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from openai.types.chat.chat_completion_message_param import ChatCompletionMessageParam
from openai.types.chat.chat_completion_user_message_param import (
    ChatCompletionUserMessageParam,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from starlette.background import BackgroundTask

from core.analyst_db import AnalystDB, get_data_source_type
from core.api import (
    AnalysisGenerationError,
    run_complete_analysis,
    summarize_conversation,
)
from core.chat_dataset_helper import cleanup_message_datasets
from core.constants import (
    ALTERNATIVE_LLM_BIG,
    CONTEXT_WARNING_THRESHOLD,
    MODEL_CONTEXT_WINDOW,
)
from core.deps import get_initialized_db
from core.logging_helper import get_logger
from core.schema import (
    AnalystChatMessage,
    AnalystDataset,
    ChatCreate,
    ChatMessagePayload,
    ChatRequest,
    ChatResponse,
    ChatUpdate,
    GetBusinessAnalysisResult,
    RunAnalysisResult,
    RunChartsResult,
    RunDatabaseAnalysisResult,
)
from core.token_tracking import (
    TiktokenCountingStrategy,
    TokenUsageTracker,
    count_messages_tokens,
)

logger = get_logger()

MAX_EXCEL_ROWS = 50000  # Maximum rows to export to Excel to prevent memory issues


async def run_complete_analysis_task(
    chat_request: ChatRequest,
    data_source: str,
    analyst_db: AnalystDB,
    chat_id: str,
    message_id: str,
    enable_chart_generation: bool,
    enable_business_insights: bool,
    request: Request,
) -> None:
    """Run the complete analysis pipeline."""
    from core.analyst_db import InternalDataSourceType

    source = get_data_source_type(data_source)
    logger.debug(
        "Running analysis for user.",
        extra={
            "data_source": data_source,
            "user_id": analyst_db.user_id,
        },
    )
    dataset_metadata = []
    if source in [InternalDataSourceType.REGISTRY, InternalDataSourceType.FILE]:
        dataset_metadata = (
            await analyst_db.list_analyst_dataset_metadata(
                InternalDataSourceType.REGISTRY
            )
        ) + (
            await analyst_db.list_analyst_dataset_metadata(InternalDataSourceType.FILE)
        )
    else:
        dataset_metadata = await analyst_db.list_analyst_dataset_metadata(source)

    run_analysis_iterator = run_complete_analysis(
        chat_request=chat_request,
        data_source=source,
        dataset_metadata=dataset_metadata,
        analyst_db=analyst_db,
        chat_id=chat_id,
        message_id=message_id,
        enable_chart_generation=enable_chart_generation,
        enable_business_insights=enable_business_insights,
        request=request,
    )

    async for message in run_analysis_iterator:
        if isinstance(message, AnalysisGenerationError):
            break
        else:
            pass


router = APIRouter(prefix="/chats", tags=["chats"])


@router.post("")
async def create_chat(
    chat: ChatCreate,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> dict[str, str]:
    """Create a new chat with optional data source."""
    chat_id = await analyst_db.create_chat(
        chat_name=chat.name,
        data_source=chat.data_source,
    )

    return {"id": chat_id}


@router.get("")
async def get_chats(
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> list[dict[str, Any]]:
    """Get all chats."""
    chat_list = await analyst_db.get_chat_list()

    return [
        {
            "id": chat["id"],
            "name": chat["name"],
            "data_source": chat.get("data_source", "catalog"),
            "created_at": chat["created_at"],
        }
        for chat in chat_list
    ]


@router.get("/{chat_id}")
async def get_chat(
    chat_id: str, analyst_db: AnalystDB = Depends(get_initialized_db)
) -> ChatResponse:
    """Get a specific chat by ID."""
    chat = await analyst_db.get_chat_messages(chat_id=chat_id)

    return {
        "id": chat_id,
        "messages": chat,
    }


@router.put("/{chat_id}")
async def update_chat(
    chat_id: str,
    chat: ChatUpdate,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> dict[str, str]:
    """Update a chat's name and/or data source."""
    response_messages = []

    # Update chat name if provided
    if chat.name:
        await analyst_db.rename_chat(chat_id, chat.name)
        response_messages.append("renamed")

    # Update data source if provided
    if chat.data_source:
        await analyst_db.update_chat_data_source(chat_id, chat.data_source)
        response_messages.append("updated data source")

    if not response_messages:
        return {"message": f"No changes made to chat with ID {chat_id}"}

    return {
        "message": f"Chat with ID {chat_id} was {' and '.join(response_messages)} successfully"
    }


@router.delete("/{chat_id}", status_code=200)
async def delete_chat(
    chat_id: str, analyst_db: AnalystDB = Depends(get_initialized_db)
) -> dict[str, str]:
    """Delete a chat."""
    # Get all messages to clean up datasets
    messages = await analyst_db.get_chat_messages(chat_id=chat_id)

    # Clean up datasets for all messages
    for message in messages:
        await cleanup_message_datasets(analyst_db, message)

    # Delete the chat
    await analyst_db.delete_chat(chat_id=chat_id)

    return {"message": f"Chat with ID {chat_id} deleted successfully"}


@router.get("/{chat_id}/messages")
async def get_chat_messages(
    chat_id: str,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> list[AnalystChatMessage]:
    """Get messages for a specific chat."""
    chat = await analyst_db.get_chat_messages(chat_id=chat_id)

    return chat


@router.delete("/messages/{message_id}")
async def delete_chat_message(
    request: Request,
    message_id: str,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> list[AnalystChatMessage]:
    """Delete a specific message."""
    try:
        message = await analyst_db.get_chat_message(message_id=message_id)
        if not message:
            raise HTTPException(
                status_code=404, detail=f"Message with ID {message_id} not found"
            )
        else:
            # Clean up associated datasets before deleting the message
            await cleanup_message_datasets(analyst_db, message)

            await analyst_db.delete_chat_message(message_id=message_id)
            messages = await analyst_db.get_chat_messages(
                chat_id=message.chat_id,
            )
            return cast(list[AnalystChatMessage], list(messages))
    except Exception as e:
        logger.error(f"Error deleting message: {str(e)}")

        return cast(list[AnalystChatMessage], [])


@router.get("/{chat_id}/messages/{message_id}")
async def get_chat_message(
    chat_id: str,
    message_id: str,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> AnalystChatMessage:
    """Get a specific message by ID from a specific chat."""
    try:
        message = await analyst_db.get_chat_message(message_id=message_id)
        if not message:
            raise HTTPException(
                status_code=404, detail=f"Message with ID {message_id} not found"
            )

        # Verify the message belongs to the specified chat
        if message.chat_id != chat_id:
            raise HTTPException(
                status_code=404,
                detail=f"Message with ID {message_id} not found in chat {chat_id}",
            )

        return message
    except Exception as e:
        logger.error(f"Error getting message: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Error retrieving message: {str(e)}"
        )


@router.post("/messages")
async def create_new_chat_message(
    request: Request,
    payload: ChatMessagePayload,
    background_tasks: BackgroundTasks,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> dict[str, Union[str, list[AnalystChatMessage], None]]:
    """Create a new chat and post a message to it."""
    # Create a new chat
    chat_id = await analyst_db.create_chat(
        chat_name=payload.chatName,
        data_source=payload.data_source,
    )

    # Create the user message
    user_message = AnalystChatMessage(
        role="user", content=payload.message, components=[]
    )

    message_id = await analyst_db.add_chat_message(
        chat_id=chat_id,
        message=user_message,
    )

    # Create valid messages for the chat request
    valid_messages: list[ChatCompletionMessageParam] = [
        user_message.to_openai_message_param()
    ]

    # Add the current message
    valid_messages.append(
        ChatCompletionUserMessageParam(role="user", content=payload.message)
    )

    # Create the chat request
    chat_request = ChatRequest(messages=valid_messages)

    # Run the analysis in the background
    background_tasks.add_task(
        run_complete_analysis_task,
        chat_request,
        payload.data_source,
        analyst_db,
        chat_id,
        message_id,
        payload.enable_chart_generation,
        payload.enable_business_insights,
        request,
    )

    chat_list = await analyst_db.get_chat_list()
    chat_name = next((n["name"] for n in chat_list if n["id"] == chat_id), None)
    messages = await analyst_db.get_chat_messages(chat_id=chat_id)

    chat = {
        "id": chat_id,
        "name": chat_name,
        "messages": messages,
    }

    return chat


@router.post("/{chat_id}/messages")
async def create_chat_message(
    request: Request,
    chat_id: str,
    payload: ChatMessagePayload,
    background_tasks: BackgroundTasks,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> dict[str, Union[str, list[AnalystChatMessage], None]]:
    """Post a message to a specific chat."""
    messages = await analyst_db.get_chat_messages(chat_id=chat_id)

    # Check if any message is in progress
    in_progress = any(message.in_progress for message in messages)

    # Check if cancelled
    if not in_progress:
        # Check if there's an existing summary (find last system message)
        last_summary_idx = None
        for i in range(len(messages) - 1, -1, -1):  # Iterate backwards through indices
            if messages[i].role == "system":
                last_summary_idx = i
                break

        # Build context: if summary exists, use [summary] + messages_after, else all messages
        if last_summary_idx is not None:
            context_messages = messages[last_summary_idx:]
            logger.info(
                f"[chat_id={chat_id}] Using summary + {len(context_messages) - 1} messages after"
            )
        else:
            context_messages = messages

        # Create valid messages for the chat request
        valid_messages: list[ChatCompletionMessageParam] = [
            msg.to_openai_message_param()
            for msg in context_messages
            if msg.content.strip()
        ]

        # Add the current message
        valid_messages.append(
            ChatCompletionUserMessageParam(role="user", content=payload.message)
        )

        # Check context usage
        tokens_used = count_messages_tokens(valid_messages, ALTERNATIVE_LLM_BIG)
        usage_pct = (tokens_used / MODEL_CONTEXT_WINDOW) * 100

        logger.info(
            f"[chat_id={chat_id}] Context: {tokens_used:,}/{MODEL_CONTEXT_WINDOW:,} tokens ({usage_pct:.1f}%)"
        )

        # Create and store the user message first (before summarization)
        user_message = AnalystChatMessage(
            role="user", content=payload.message, components=[]
        )
        message_id = await analyst_db.add_chat_message(
            chat_id=chat_id, message=user_message
        )
        user_message.id = message_id

        # Trigger summarization if over threshold
        if tokens_used >= CONTEXT_WARNING_THRESHOLD:
            logger.warning(
                f"[chat_id={chat_id}] ⚠️  Context at {usage_pct:.1f}% - creating summary"
            )

            # Create new system message with in_progress=True
            summary_message = AnalystChatMessage(
                role="system",
                content="Summarizing conversation...",
                components=[],
                in_progress=True,
            )
            summary_message.id = await analyst_db.add_chat_message(
                chat_id=chat_id, message=summary_message
            )

            # Determine what to summarize
            if last_summary_idx is not None:
                # Summarize from last system message onwards
                messages_to_summarize = [
                    msg.to_openai_message_param()
                    for msg in messages[last_summary_idx:]
                    if msg.content.strip()
                ]
            else:
                # Summarize all messages
                messages_to_summarize = [
                    msg.to_openai_message_param()
                    for msg in messages
                    if msg.content.strip()
                ]

            # Create token tracker for summarization
            summarization_tracker = TokenUsageTracker(
                strategy=TiktokenCountingStrategy()
            )

            try:
                summary_text = await summarize_conversation(
                    messages_to_summarize, token_tracker=summarization_tracker
                )

                logger.info(
                    f"[chat_id={chat_id}] Summarization token usage: "
                    f"{summarization_tracker.prompt_tokens} prompt + "
                    f"{summarization_tracker.completion_tokens} completion = "
                    f"{summarization_tracker.total_tokens} total tokens"
                )

                # Update the summary message with actual content
                summary_message.content = summary_text
                summary_message.in_progress = False

                await analyst_db.update_chat_message(
                    message_id=summary_message.id,
                    message=summary_message,
                )

                logger.info(
                    f"[chat_id={chat_id}] Summary stored ({len(summary_text)} chars)"
                )

                # Rebuild context using the new summary
                summary_param = summary_message.to_openai_message_param()
                valid_messages = [
                    summary_param,
                    valid_messages[-1],
                ]  # summary + current message

            except Exception as e:
                logger.error(f"[chat_id={chat_id}] Failed to create summary: {e}")
                # Mark summary as failed
                summary_message.content = "Failed to create summary"
                summary_message.in_progress = False
                summary_message.error = str(e)
                await analyst_db.update_chat_message(
                    message_id=summary_message.id,
                    message=summary_message,
                )
                # Continue without summary

        # Create the chat request
        chat_request = ChatRequest(messages=valid_messages)

        # Run the analysis in the background
        background_tasks.add_task(
            run_complete_analysis_task,
            chat_request,
            payload.data_source,
            analyst_db,
            chat_id,
            message_id,
            payload.enable_chart_generation,
            payload.enable_business_insights,
            request,
        )

    chat_list = await analyst_db.get_chat_list()
    chat_name = next((n["name"] for n in chat_list if n["id"] == chat_id), None)
    messages = await analyst_db.get_chat_messages(chat_id=chat_id)

    chat = {
        "id": chat_id,
        "name": chat_name,
        "messages": messages,
    }

    return chat


@router.get("/{chat_id}/messages/download/")
async def save_chat_messages(
    request: Request,
    chat_id: str,
    analyst_db: AnalystDB = Depends(get_initialized_db),
    message_id: str | None = None,
) -> StreamingResponse:
    """Download chat messages as an Excel spreadsheet."""
    temp_files: list[str] = []

    chat_messages: List[AnalystChatMessage] = await analyst_db.get_chat_messages(
        chat_id=chat_id
    )

    # If a specific message_id is provided, filter messages
    if message_id:
        idx = next((i for i, m in enumerate(chat_messages) if m.id == message_id), None)
        if idx is None or chat_messages[idx].role != "user":
            raise HTTPException(detail="User message not found", status_code=404)

        # Find the following assistant message
        assistant_message = None
        for i in range(idx + 1, len(chat_messages)):
            if chat_messages[i].role == "assistant":
                assistant_message = chat_messages[i]
                break

        # Include user message and assistant response if found
        if assistant_message:
            filtered_messages = [chat_messages[idx], assistant_message]
        else:
            filtered_messages = [chat_messages[idx]]

        chat_messages = filtered_messages

    if not chat_messages:
        # Create an empty workbook for empty chats
        analysis_workbook = Workbook()
        analysis_workbook.remove(analysis_workbook.active)

        empty_sheet = analysis_workbook.create_sheet("Empty Chat")
        empty_sheet["A1"] = "Chat Export"
        empty_sheet["A3"] = "This chat contains no messages to export yet."

        output = io.BytesIO()
        analysis_workbook.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if any(msg.in_progress for msg in chat_messages):
        raise HTTPException(
            detail="Cannot download while a chat is in progress.", status_code=425
        )

    analysis_workbook = Workbook()
    data_sheets_count = 0
    charts_sheets_count = 0
    report_sheets_count = 0

    # Remove the initial default sheet created by Workbook()
    analysis_workbook.remove(analysis_workbook.active)

    for i, chat_message in enumerate(chat_messages):
        # Skip system messages (summarization messages)
        if chat_message.role == "system":
            continue

        if chat_message.role == "assistant":
            # Handle Analysis Report sheet
            report_sheets_count += 1
            report_sheet_name = (
                "Sheet" if report_sheets_count == 1 else f"Sheet {report_sheets_count}"
            )
            report_sheet = analysis_workbook.create_sheet(report_sheet_name)

            report_sheet["A1"] = "Analysis Report"

            # Find the previous user message by searching backwards
            user_question = "No question found"
            for j in range(i - 1, -1, -1):
                if chat_messages[j].role == "user":
                    user_question = chat_messages[j].content
                    break

            report_sheet["A3"] = "Question"
            report_sheet["A4"] = user_question
            report_sheet["A6"] = "Answer"
            report_sheet["A7"] = chat_message.content

            business_components: list[GetBusinessAnalysisResult] = [
                component
                for component in chat_message.components
                if isinstance(component, GetBusinessAnalysisResult)
            ]
            for index, business_component in enumerate(business_components):
                cell_index = index + 1  # Excel uses 1 indexing
                report_sheet.cell(9, cell_index).value = "Bottom Line"
                report_sheet.cell(10, cell_index).value = business_component.bottom_line

                report_sheet.cell(12, cell_index).value = "Additional Insights"
                report_sheet.cell(
                    13, cell_index
                ).value = business_component.additional_insights

                report_sheet.cell(15, cell_index).value = "Follow-up Questions:"
                for q_index, followup_question in enumerate(
                    business_component.follow_up_questions
                ):
                    report_sheet.cell(
                        16 + q_index, cell_index
                    ).value = followup_question

            # Handle Data sheets
            run_analysis_components: List[
                RunAnalysisResult | RunDatabaseAnalysisResult
            ] = [
                component
                for component in chat_message.components
                if isinstance(component, (RunAnalysisResult, RunDatabaseAnalysisResult))
            ]
            for run_analysis_component in run_analysis_components:
                # Load dataset from storage
                if not run_analysis_component.dataset_id:
                    continue

                try:
                    from core.analyst_db import DatasetType

                    df = await analyst_db.dataset_handler.get_dataframe(
                        run_analysis_component.dataset_id,
                        expected_type=DatasetType.ANALYST_RESULT_DATASET,
                        max_rows=None,
                    )
                    metadata = await analyst_db.dataset_handler.get_dataset_metadata(
                        run_analysis_component.dataset_id
                    )
                    dataset_to_export = AnalystDataset(
                        name=metadata.original_name,
                        data=df,
                    )
                except Exception as e:
                    logger.error(
                        f"Failed to load dataset {run_analysis_component.dataset_id}: {e}"
                    )
                    continue

                data_sheets_count += 1
                data_sheet_name = (
                    "Data" if data_sheets_count == 1 else f"Data {data_sheets_count}"
                )
                data_sheet = analysis_workbook.create_sheet(data_sheet_name)

                try:
                    dataset: polars.dataframe.frame.DataFrame = (
                        dataset_to_export.data.df
                    )

                    # Convert to pandas with error handling for large datasets
                    pandas_df = dataset.to_pandas()

                    # Add size check to prevent memory issues and Excel limits
                    original_rows = pandas_df.shape[0]
                    if original_rows > MAX_EXCEL_ROWS:
                        logger.warning(
                            f"Dataset too large ({original_rows} rows), truncating to {MAX_EXCEL_ROWS} rows"
                        )
                        pandas_df = pandas_df.head(MAX_EXCEL_ROWS)
                        # Add a notice row at the top of the sheet
                        data_sheet.append(
                            [
                                f"NOTICE: Dataset truncated from {original_rows:,} to {MAX_EXCEL_ROWS:,} rows due to Excel limitations"
                            ]
                        )
                        data_sheet.append([])

                    for r in dataframe_to_rows(pandas_df, index=False, header=True):
                        data_sheet.append(r)

                except Exception as e:
                    logger.error(f"Failed to process dataset: {e}")
                    # Create error sheet instead of crashing
                    data_sheet["A1"] = "Dataset Processing Error"
                    data_sheet["A2"] = f"Error: {str(e)}"

            # Handle Charts sheets
            run_charts_components: List[RunChartsResult] = [
                component
                for component in chat_message.components
                if isinstance(component, RunChartsResult)
            ]
            for run_chart_component in run_charts_components:
                for f, js in [
                    (run_chart_component.fig1, run_chart_component.fig1_json),
                    (run_chart_component.fig2, run_chart_component.fig2_json),
                ]:
                    if not js or not f:
                        continue
                    charts_sheets_count += 1
                    charts_sheet_name = f"Chart {charts_sheets_count}"
                    charts_sheet = analysis_workbook.create_sheet(charts_sheet_name)

                    # Save the chart data
                    try:
                        parsed_json = json.loads(js)
                        data_list = parsed_json.get("data", [])
                        if data_list and len(data_list) > 0:
                            fig_json = data_list[0].copy()
                            [fig_json.pop(k, None) for k in ["marker", "name", "type"]]
                            chart_df = pd.DataFrame(fig_json)
                            for r in dataframe_to_rows(
                                chart_df, index=False, header=True
                            ):
                                charts_sheet.append(r)

                        with tempfile.NamedTemporaryFile(
                            suffix=".png", delete=False
                        ) as tmpfile:
                            f.write_image(tmpfile.name)
                            img = XLImage(tmpfile.name)
                            charts_sheet.add_image(img, "F3")
                            temp_files.append(tmpfile.name)
                    except (
                        json.JSONDecodeError,
                        KeyError,
                        IndexError,
                        ValueError,
                    ) as e:
                        logger.warning(f"Failed to process chart data: {e}")
                        # Create error sheet instead of crashing
                        charts_sheet["A1"] = "Chart Processing Error"
                        charts_sheet["A2"] = f"Error: {str(e)}"
                    except Exception as e:
                        logger.error(f"Unexpected error processing chart: {e}")
                        continue  # Skip this chart but continue processing

    output = io.BytesIO()
    analysis_workbook.save(output)
    output.seek(0)

    # Create background task to cleanup temporary files
    def cleanup_files(file_paths: List[str]) -> None:
        for fp in file_paths:
            if os.path.exists(fp):
                os.remove(fp)

    background_task = BackgroundTask(cleanup_files, temp_files) if temp_files else None

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=background_task,
    )
