# Copyright 2025 DataRobot, Inc.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import io
import json
import os
import pytest
from typing import Any, Generator
from unittest.mock import AsyncMock, MagicMock, patch
from unittest.mock import MagicMock as MagicMockType

import polars as pl
from pytest import MonkeyPatch
from fastapi import Request, Response
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import core.deps
from core.analyst_db import (
    AnalystDB,  # noqa: E402
    DatasetType,
)
from core.datarobot_client import use_user_token
from core.file_utils import (
    detect_and_decode_csv,
    detect_delimiter,
    load_and_validate_csv,
)
from core.middleware import (
    SessionState,
    _initialize_session,
    _set_session_cookie,
)
from core.rest_api import create_app
from core.routers.chats import delete_chat_message
from core.routers.user import get_datarobot_account, store_datarobot_account
from core.schema import (
    AnalystChatMessage,
    GetBusinessAnalysisMetadata,
    GetBusinessAnalysisResult,
    RunAnalysisResult,
    RunAnalysisResultMetadata,
    RunChartsResult,
)


@pytest.fixture
def mock_analyst_db() -> AsyncMock:
    return AsyncMock(spec=AnalystDB)


@pytest.fixture
def test_client() -> Generator[TestClient, None, None]:
    app = create_app()
    client = TestClient(app)
    yield client


@pytest.fixture
def mock_session_state() -> SessionState:
    session = SessionState()
    session._state = {
        "datarobot_account_info": None,
        "datarobot_api_scoped_token": None,
        "analyst_db": None,
    }
    return session


@pytest.fixture
def mock_request() -> MagicMockType:
    request = MagicMock(spec=Request)
    request.state = MagicMock()
    request.cookies = {}
    request.headers = {}
    request.method = "GET"
    return request

@pytest.fixture
def clean_environ(monkeypatch: pytest.MonkeyPatch) -> Generator[None, None, None]:
    monkeypatch.delenv("TEST_USER_EMAIL", raising=False)
    yield


@pytest.mark.usefixtures("clean_environ")
def test_initialize_session_default_values(mock_request: MagicMockType) -> None:
    """Test that _initialize_session creates a new session with default values"""

    async def run_test() -> None:
        session_state, session_id, user_id = await _initialize_session(mock_request)
        assert session_state is not None
        assert session_id is None
        assert user_id is None
        assert session_state._state["datarobot_account_info"] is None
        assert session_state._state["datarobot_api_scoped_token"] is None
        assert session_state._state["analyst_db"] is None

    import asyncio

    asyncio.run(run_test())


@pytest.mark.usefixtures("clean_environ")
def test_set_session_cookie() -> None:
    """Test that _set_session_cookie sets cookies correctly"""
    response = Response()
    user_id = "test_user_id"
    session_id = "test_session_id"

    # Test when user_id exists but no cookie
    _set_session_cookie(response, user_id, session_id, None)

    # The cookie should be set with base64 encoded user_id
    cookies = [
        header for header in response.raw_headers if header[0].decode() == "set-cookie"
    ]
    assert cookies, "No set-cookie header found"
    assert "session_fastapi" in cookies[0][1].decode()

    # Test when neither user_id nor cookie exists
    response = Response()
    _set_session_cookie(response, None, session_id, None)

    # The cookie should be set with session_id
    cookies = [
        header for header in response.raw_headers if header[0].decode() == "set-cookie"
    ]
    assert cookies, "No set-cookie header found"
    assert "session_fastapi" in cookies[0][1].decode()

    # Test when cookie already exists
    response = Response()
    _set_session_cookie(response, user_id, session_id, "existing_cookie")

    # No cookie should be set
    cookies = [
        header for header in response.raw_headers if header[0].decode() == "set-cookie"
    ]
    assert not cookies, "Cookie was set when it shouldn't have been"


@pytest.mark.usefixtures("clean_environ")
@pytest.mark.asyncio
async def test_use_user_token_from_request_header(mock_request: MagicMockType) -> None:
    """Test that use_user_token uses the scoped token from request header by default"""
    # Setup mock request with a session that has a scoped token
    mock_request.state.session = MagicMock()
    mock_request.state.session.datarobot_api_scoped_token = None
    mock_request.headers = {"x-datarobot-api-key": "visitor_token"}

    # Mock dr.Client to track which token is used
    with (
        patch("datarobot.Client") as mock_dr_client,
        patch.dict(
            os.environ,
            {"DATAROBOT_ENDPOINT": "https://app.datarobot.com/api/v2"},
            clear=True,
        ),
    ):
        # Use the context manager
        with use_user_token(mock_request):
            pass

        # Check that dr.Client was called with the scoped token
        mock_dr_client.assert_called_once_with(
            token="visitor_token", endpoint="https://app.datarobot.com/api/v2"
        )


@pytest.mark.asyncio
async def test_use_user_token_from_session_if_set(mock_request: MagicMockType) -> None:
    """Test that use_user_token uses token form session if present"""
    # Setup mock request with a session that has only a regular token
    mock_request.state.session = MagicMock()
    mock_request.state.session.datarobot_api_scoped_token = "token_in_session"
    mock_request.headers = {"x-datarobot-api-key": "visitor_token"}

    # Mock dr.Client to track which token is used
    with (
        patch.dict(
            os.environ,
            {"DATAROBOT_ENDPOINT": "https://app.datarobot.com/api/v2"},
            clear=True,
        ),
        patch("datarobot.Client") as mock_dr_client,
    ):
        # Use the context manager
        with use_user_token(mock_request):
            pass

        # Check that dr.Client was called with the regular token
        mock_dr_client.assert_called_once_with(
            token="token_in_session", endpoint="https://app.datarobot.com/api/v2"
        )


@pytest.mark.usefixtures("clean_environ")
@pytest.mark.asyncio
async def test_get_datarobot_account_includes_scoped_token(
    mock_request: MagicMockType,
) -> None:
    """Test that get_datarobot_account includes the scoped token in response"""
    # Setup mock request with session containing all tokens
    mock_request.state.session = MagicMock()
    mock_request.state.session.datarobot_account_info = {"uid": "test_uid"}
    mock_request.state.session.datarobot_api_scoped_token = None
    mock_request.headers = {"x-datarobot-api-key": "visitor_token_987654321"}

    # Call the endpoint function directly
    with patch.dict(
        os.environ, {"DATAROBOT_API_TOKEN": "creator_token_123456789"}, clear=True
    ):
        response = await get_datarobot_account(mock_request)

    # Check that response includes all expected fields
    assert response["datarobot_account_info"] == {"uid": "test_uid"}
    assert response["datarobot_api_token"] == "****6789"
    assert response["datarobot_api_scoped_token"] == "****4321"


@pytest.mark.usefixtures("clean_environ")
@pytest.mark.asyncio
async def test_store_datarobot_account_handles_scoped_token(
    mock_request: MagicMockType,
    mock_session_state: SessionState,
) -> None:
    """Test that store_datarobot_account can save the scoped token"""
    # Setup mock request
    mock_request.state.session = mock_session_state

    # Only test the keys that are supported
    mock_json = AsyncMock(return_value={"api_token": "api_token_123"})
    mock_request.json = mock_json

    # Call the endpoint function directly
    response = await store_datarobot_account(mock_request)

    # Check that tokens were saved to session
    assert mock_request.state.session.datarobot_api_scoped_token == "api_token_123"
    assert response["success"] is True


@pytest.mark.usefixtures("clean_environ")
@pytest.mark.asyncio
async def test_delete_chat_message(
    mock_request: MagicMockType,
    mock_analyst_db: AsyncMock,
) -> None:
    mock_request.state.session = MagicMock()
    mock_analyst_db.chat_handler = AsyncMock()

    # Set up the mocks for the new function signature
    mock_message = AnalystChatMessage(
        id="message_id",
        role="user",
        content="User message 1",
        components=[],
        chat_id="test_chat_id",
    )
    mock_analyst_db.get_chat_message.return_value = mock_message
    mock_analyst_db.delete_chat_message.return_value = True

    test_messages = [
        AnalystChatMessage(role="user", content="User message 1", components=[]),
        AnalystChatMessage(
            role="assistant", content="Assistant message 1", components=[]
        ),
        AnalystChatMessage(role="user", content="User message 2", components=[]),
        AnalystChatMessage(
            role="assistant", content="Assistant message 2", components=[]
        ),
    ]
    mock_analyst_db.get_chat_messages.return_value = test_messages.copy()

    # Test successful deletion
    result = await delete_chat_message(mock_request, "message_id", mock_analyst_db)

    # Verify the mock interactions and results
    mock_analyst_db.get_chat_message.assert_called_once_with(message_id="message_id")
    mock_analyst_db.delete_chat_message.assert_called_once_with(message_id="message_id")
    mock_analyst_db.get_chat_messages.assert_called_once_with(chat_id="test_chat_id")
    assert result == test_messages.copy()

    # Reset mocks for next test
    mock_analyst_db.get_chat_message.reset_mock()
    mock_analyst_db.delete_chat_message.reset_mock()
    mock_analyst_db.get_chat_messages.reset_mock()

    # Test when message not found
    mock_analyst_db.get_chat_message.return_value = None

    result = await delete_chat_message(
        mock_request, "missing_message_id", mock_analyst_db
    )

    # Should return an empty list when message not found
    assert isinstance(result, list)
    assert len(result) == 0
    mock_analyst_db.delete_chat_message.assert_not_called()

    # Reset mocks for next test
    mock_analyst_db.get_chat_message.reset_mock()
    mock_analyst_db.delete_chat_message.reset_mock()

    # Test when an exception occurs
    mock_analyst_db.get_chat_message.return_value = mock_message
    mock_analyst_db.delete_chat_message.side_effect = Exception("Test exception")

    result = await delete_chat_message(mock_request, "message_id", mock_analyst_db)

    # Should return an empty list when an exception occurs
    assert isinstance(result, list)
    assert len(result) == 0


@pytest.mark.usefixtures("clean_environ")
@pytest.mark.asyncio
async def test_save_chat_history_to_xlsx(
    example_chat_file_content: list[dict[str, Any]],
    mock_analyst_db_creation: None,
) -> None:
    user_name = "data_analyst_app_user@datarobot.com"
    # Make a request with Authorization header
    chat_history = [
        AnalystChatMessage.model_validate(chat) for chat in example_chat_file_content
    ]
    mock_analyst_db = AsyncMock()
    mock_analyst_db.get_chat_messages = AsyncMock(return_value=chat_history)
    app = create_app()
    app.dependency_overrides[core.deps.get_initialized_db] = lambda: mock_analyst_db
    client = TestClient(app)

    response = client.get(
        "/api/v1/chats/chat_id_123/messages/download/",
        headers={
            "Authorization": "Bearer test_token_123",
            "x-user-email": user_name,
        },
    )
    assert response.status_code == 200, response.json()
    wb = load_workbook(filename=io.BytesIO(response.content), read_only=True)
    # Just verify we're getting data back
    assert wb


@pytest.mark.usefixtures("clean_environ")
@pytest.mark.asyncio
async def test_save_empty_chat_to_xlsx(mock_analyst_db_creation: None) -> None:
    """Test Excel export for empty chat creates proper empty workbook"""
    user_name = "data_analyst_app_user@datarobot.com"

    mock_analyst_db = AsyncMock()
    mock_analyst_db.get_chat_messages = AsyncMock(return_value=[])
    app = create_app()
    app.dependency_overrides[core.deps.get_initialized_db] = lambda: mock_analyst_db
    client = TestClient(app)

    response = client.get(
        "/api/v1/chats/empty_chat_123/messages/download/",
        headers={
            "Authorization": "Bearer test_token_123",
            "x-user-email": user_name,
        },
    )

    assert response.status_code == 200

    # Validate workbook structure for empty chat
    wb = load_workbook(filename=io.BytesIO(response.content), read_only=True)
    sheet_names = wb.sheetnames
    assert "Empty Chat" in sheet_names

    empty_sheet = wb["Empty Chat"]
    assert empty_sheet["A1"].value == "Chat Export"
    assert "no messages to export" in empty_sheet["A3"].value.lower()


@pytest.mark.usefixtures("clean_environ")
@pytest.mark.asyncio
async def test_save_chat_with_in_progress_message_blocked(
    mock_analyst_db_creation: None,
) -> None:
    """Test that Excel export is blocked when chat has in-progress messages"""
    user_name = "data_analyst_app_user@datarobot.com"

    # Create chat history with one message in progress
    chat_history = [
        AnalystChatMessage(role="user", content="Test question", components=[]),
        AnalystChatMessage(
            role="assistant", content="", components=[], in_progress=True
        ),
    ]

    mock_analyst_db = AsyncMock()
    mock_analyst_db.get_chat_messages = AsyncMock(return_value=chat_history)
    app = create_app()
    app.dependency_overrides[core.deps.get_initialized_db] = lambda: mock_analyst_db
    client = TestClient(app)

    response = client.get(
        "/api/v1/chats/chat_in_progress/messages/download/",
        headers={
            "Authorization": "Bearer test_token_123",
            "x-user-email": user_name,
        },
    )

    assert response.status_code == 425
    assert (
        "cannot download while a chat is in progress"
        in response.json()["detail"].lower()
    )


@pytest.mark.usefixtures("clean_environ")
@pytest.mark.asyncio
async def test_save_single_message_to_xlsx(mock_analyst_db_creation: None) -> None:
    """Test Excel export for single message with message_id parameter"""
    user_name = "data_analyst_app_user@datarobot.com"

    # Create multiple messages but we'll only export one pair
    chat_history = [
        AnalystChatMessage(
            id="msg1",
            role="user",
            content="First question",
            components=[],
            in_progress=False,
        ),
        AnalystChatMessage(
            id="msg2",
            role="assistant",
            content="First answer",
            components=[],
            in_progress=False,
        ),
        AnalystChatMessage(
            id="msg3",
            role="user",
            content="Second question",
            components=[],
            in_progress=False,
        ),
        AnalystChatMessage(
            id="msg4",
            role="assistant",
            content="Second answer",
            components=[],
            in_progress=False,
        ),
    ]

    mock_analyst_db = AsyncMock()
    mock_analyst_db.get_chat_messages = AsyncMock(return_value=chat_history)
    app = create_app()
    app.dependency_overrides[core.deps.get_initialized_db] = lambda: mock_analyst_db
    client = TestClient(app)

    response = client.get(
        "/api/v1/chats/chat_123/messages/download/?message_id=msg3",
        headers={
            "Authorization": "Bearer test_token_123",
            "x-user-email": user_name,
        },
    )

    assert response.status_code == 200

    # Validate workbook content
    wb = load_workbook(filename=io.BytesIO(response.content), read_only=True)
    assert wb

    # Should have analysis sheet for the single message pair
    sheet_names = wb.sheetnames
    assert "Sheet" in sheet_names

    report_sheet = wb["Sheet"]
    assert report_sheet["A1"].value == "Analysis Report"
    assert report_sheet["A4"].value == "Second question"  # User question
    assert report_sheet["A7"].value == "Second answer"  # Assistant answer


@pytest.mark.usefixtures("clean_environ")
@pytest.mark.asyncio
async def test_export_message_to_xlsx(
    mock_analyst_db_creation: None,
) -> None:
    """
    Test comprehensive Excel export for individual message with dataset_id and charts.
    Validates that report sheet, data sheet, and chart sheets are all properly exported.
    """

    user_name = "data_analyst_app_user@datarobot.com"

    # Create a sample dataframe for the dataset
    sample_data = pl.DataFrame(
        {
            "category": ["A", "B", "C"],
            "value": [10, 20, 30],
            "score": [85.5, 92.3, 78.9],
        }
    )

    # Create plotly figure JSONs (bar chart and scatter plot)
    fig1_json = json.dumps(
        {
            "data": [
                {
                    "x": ["A", "B", "C"],
                    "y": [10, 20, 30],
                    "type": "bar",
                    "name": "Values",
                    "marker": {"color": "blue"},
                }
            ],
            "layout": {"title": "Bar Chart"},
        }
    )

    fig2_json = json.dumps(
        {
            "data": [
                {
                    "x": ["A", "B", "C"],
                    "y": [85.5, 92.3, 78.9],
                    "type": "scatter",
                    "mode": "lines+markers",
                    "name": "Scores",
                    "marker": {"size": 8},
                }
            ],
            "layout": {"title": "Score Trend"},
        }
    )

    # Create messages with both dataset and charts components
    chat_history = [
        AnalystChatMessage(
            id="msg1",
            role="user",
            content="Analyze the data and show visualizations",
            components=[],
            in_progress=False,
        ),
        # Include a system message to test that it gets skipped during export
        AnalystChatMessage(
            id="system1",
            role="system",
            content="Summary: User requested data analysis and visualizations",
            components=[],
            in_progress=False,
        ),
        AnalystChatMessage(
            id="msg2",
            role="assistant",
            content="Here is the complete analysis with data and charts",
            components=[
                RunAnalysisResult(
                    status="success",
                    code="df = analyze_data(data)",
                    dataset=None,  # Dataset is not loaded in memory
                    dataset_id="dataset_123",  # Reference to stored dataset
                    metadata=RunAnalysisResultMetadata(
                        duration=1.5,
                        attempts=1,
                        datasets_analyzed=1,
                        total_rows_analyzed=3,
                        total_columns_analyzed=3,
                    ),
                ),
                RunChartsResult(
                    status="success",
                    fig1_json=fig1_json,
                    fig2_json=fig2_json,
                    code="create_charts(df)",
                    metadata=RunAnalysisResultMetadata(
                        duration=0.5,
                        attempts=1,
                        datasets_analyzed=1,
                        total_rows_analyzed=3,
                        total_columns_analyzed=3,
                    ),
                ),
                GetBusinessAnalysisResult(
                    status="success",
                    bottom_line="The analysis shows strong correlation between categories and values",
                    additional_insights="Category A has the lowest value while Category C has the highest value, indicating a positive trend",
                    follow_up_questions=[
                        "What factors drive the value differences?",
                        "Are there seasonal patterns in the data?",
                        "How do these trends compare to industry benchmarks?",
                    ],
                    metadata=GetBusinessAnalysisMetadata(
                        duration=0.3,
                        rows_analyzed=3,
                        columns_analyzed=3,
                    ),
                ),
            ],
            in_progress=False,
        ),
    ]

    # Create mock dataset handler
    mock_dataset_handler = AsyncMock()
    mock_dataset_handler.get_dataframe = AsyncMock(return_value=sample_data)
    mock_dataset_metadata = MagicMock()
    mock_dataset_metadata.original_name = "test_dataset"
    mock_dataset_handler.get_dataset_metadata = AsyncMock(
        return_value=mock_dataset_metadata
    )

    # Create mock analyst_db
    mock_analyst_db = AsyncMock()
    mock_analyst_db.get_chat_messages = AsyncMock(return_value=chat_history)
    mock_analyst_db.dataset_handler = mock_dataset_handler

    app = create_app()
    app.dependency_overrides[core.deps.get_initialized_db] = lambda: mock_analyst_db
    client = TestClient(app)

    response = client.get(
        "/api/v1/chats/chat_123/messages/download/?message_id=msg1",
        headers={
            "Authorization": "Bearer test_token_123",
            "x-user-email": user_name,
        },
    )

    assert response.status_code == 200

    # Validate workbook content
    wb = load_workbook(filename=io.BytesIO(response.content), read_only=True)
    assert wb

    sheet_names = wb.sheetnames
    # Should have all types: Sheet (report), Data sheet, and Chart sheets
    assert "Sheet" in sheet_names, "Missing report sheet"
    assert "Data" in sheet_names, "Missing data sheet"
    assert "Chart 1" in sheet_names, "Missing chart 1 sheet"
    assert "Chart 2" in sheet_names, "Missing chart 2 sheet"

    # Verify system message was skipped (only 1 report sheet for 1 assistant message)
    report_sheets = [s for s in sheet_names if s.startswith("Sheet")]
    assert len(report_sheets) == 1, (
        f"Expected 1 report sheet (system message should be skipped), got {len(report_sheets)}"
    )

    # Verify report sheet (analysis summary) - comprehensive validation
    report_sheet = wb["Sheet"]
    assert report_sheet["A1"].value == "Analysis Report"

    # Question section
    assert report_sheet["A3"].value == "Question"
    assert report_sheet["A4"].value == "Analyze the data and show visualizations"

    # Answer section
    assert report_sheet["A6"].value == "Answer"
    assert (
        report_sheet["A7"].value == "Here is the complete analysis with data and charts"
    )

    # Business insights - Bottom Line
    assert report_sheet["A9"].value == "Bottom Line"
    assert "strong correlation" in report_sheet["A10"].value.lower(), (
        "Bottom line content missing"
    )

    # Business insights - Additional Insights
    assert report_sheet["A12"].value == "Additional Insights"
    assert "category a" in report_sheet["A13"].value.lower(), (
        "Additional insights content missing"
    )

    # Follow-up Questions
    assert report_sheet["A15"].value == "Follow-up Questions:"
    assert "factors" in report_sheet["A16"].value.lower()
    assert "seasonal" in report_sheet["A17"].value.lower()
    assert "benchmarks" in report_sheet["A18"].value.lower()

    # Verify data sheet was created with the loaded dataset
    data_sheet = wb["Data"]
    # Check headers
    assert data_sheet["A1"].value == "category"
    assert data_sheet["B1"].value == "value"
    assert data_sheet["C1"].value == "score"
    # Check first data row
    assert data_sheet["A2"].value == "A"
    assert data_sheet["B2"].value == 10
    assert data_sheet["C2"].value == 85.5
    # Check second data row
    assert data_sheet["A3"].value == "B"
    assert data_sheet["B3"].value == 20
    assert data_sheet["C3"].value == 92.3

    # Verify chart 1 sheet (bar chart data)
    chart1_sheet = wb["Chart 1"]
    assert chart1_sheet["A1"].value == "x"
    assert chart1_sheet["B1"].value == "y"
    assert chart1_sheet["A2"].value == "A"
    assert chart1_sheet["B2"].value == 10
    assert chart1_sheet["A3"].value == "B"
    assert chart1_sheet["B3"].value == 20
    assert chart1_sheet["A4"].value == "C"
    assert chart1_sheet["B4"].value == 30

    # Verify chart 2 sheet (scatter plot data)
    chart2_sheet = wb["Chart 2"]
    assert chart2_sheet["A1"].value == "x"
    assert chart2_sheet["B1"].value == "y"
    assert chart2_sheet["A2"].value == "A"
    assert chart2_sheet["B2"].value == 85.5
    assert chart2_sheet["A3"].value == "B"
    assert chart2_sheet["B3"].value == 92.3
    assert chart2_sheet["A4"].value == "C"
    assert chart2_sheet["B4"].value == 78.9

    # Verify the dataset handler methods were called correctly
    mock_dataset_handler.get_dataframe.assert_called_once_with(
        "dataset_123",
        expected_type=DatasetType.ANALYST_RESULT_DATASET,
        max_rows=None,
    )
    mock_dataset_handler.get_dataset_metadata.assert_called_once_with("dataset_123")


@pytest.mark.usefixtures("clean_environ")
def test_download_dictionary_with_bom(test_client: TestClient) -> None:
    class FakeDF:
        def write_csv(self, s: io.StringIO) -> None:
            s.write("column\n値")

    class FakeDictionary:
        def __init__(self, name: str) -> None:
            self.name = name

        def to_application_df(self) -> FakeDF:
            return FakeDF()

    mock_analyst_db = AsyncMock()
    mock_analyst_db.get_data_dictionary = AsyncMock(
        return_value=FakeDictionary("ja_ds")
    )
    app = create_app()
    app.dependency_overrides[core.deps.get_initialized_db] = lambda: mock_analyst_db
    client = TestClient(app)

    response = client.get("/api/v1/dictionaries/ja_ds/download?bom=true")
    assert response.status_code == 200
    # UTF-8 BOM prefix
    assert response.content.startswith(b"\xef\xbb\xbf")


@pytest.mark.usefixtures("clean_environ")
@pytest.mark.asyncio
async def test_save_chat_filename_generation(mock_analyst_db_creation: None) -> None:
    """Test proper filename generation for different scenarios"""
    user_name = "data_analyst_app_user@datarobot.com"

    chat_history = [
        AnalystChatMessage(
            id="test_msg", role="user", content="Test", components=[], in_progress=False
        ),
        AnalystChatMessage(
            role="assistant", content="Response", components=[], in_progress=False
        ),
    ]

    mock_analyst_db = AsyncMock()
    mock_analyst_db.get_chat_messages = AsyncMock(return_value=chat_history)
    app = create_app()
    app.dependency_overrides[core.deps.get_initialized_db] = lambda: mock_analyst_db
    client = TestClient(app)

    # Test full chat download filename
    response = client.get(
        "/api/v1/chats/my_chat_456/messages/download/",
        headers={
            "Authorization": "Bearer test_token_123",
            "x-user-email": user_name,
        },
    )

    assert response.status_code == 200

    # Test single message download filename
    response = client.get(
        "/api/v1/chats/my_chat_456/messages/download/?message_id=test_msg",
        headers={
            "Authorization": "Bearer test_token_123",
            "x-user-email": user_name,
        },
    )

    assert response.status_code == 200


@pytest.mark.usefixtures("clean_environ")
@pytest.mark.asyncio
async def test_download_result_dataset(test_client: TestClient) -> None:
    """Test dataset download with CSV format, BOM support, and error handling"""
    dataset_id = "a1b2c3d4-5678-90ab-cdef-1234567890ab"

    # Create test dataframe with special characters to test BOM
    test_df = pl.DataFrame(
        {
            "name": ["Alice", "Bob", "Golda"],
            "age": [25, 30, 35],
            "city": ["NYC", "München", "Zürich"],
        }
    )

    mock_dataset_handler = AsyncMock()
    mock_dataset_handler.get_dataframe = AsyncMock(return_value=test_df)

    mock_analyst_db = AsyncMock()
    mock_analyst_db.dataset_handler = mock_dataset_handler

    app = create_app()
    app.dependency_overrides[core.deps.get_initialized_db] = lambda: mock_analyst_db
    client = TestClient(app)

    # Basic download without BOM
    response = client.get(f"/api/v1/datasets/{dataset_id}/download")

    assert response.status_code == 200
    assert response.headers["Content-Type"] == "text/csv; charset=utf-8"

    content_disposition = response.headers["Content-Disposition"]
    assert "attachment" in content_disposition
    assert f"analysis_result_dataset_{dataset_id[:8]}.csv" in content_disposition

    csv_text = response.text
    assert not csv_text.startswith("\ufeff")  # No BOM
    assert "name,age,city" in csv_text
    assert "Alice,25,NYC" in csv_text
    assert "Golda" in csv_text

    # Download with BOM
    response_with_bom = client.get(f"/api/v1/datasets/{dataset_id}/download?bom=true")

    assert response_with_bom.status_code == 200
    csv_with_bom = response_with_bom.text
    assert csv_with_bom.startswith("\ufeff")  # Has BOM
    assert "name,age,city" in csv_with_bom

    # Verify mock was called with correct parameters
    assert mock_dataset_handler.get_dataframe.call_count == 2
    mock_dataset_handler.get_dataframe.assert_called_with(
        dataset_id,
        expected_type=DatasetType.ANALYST_RESULT_DATASET,
        max_rows=None,
    )

    # Dataset not found
    mock_dataset_handler.get_dataframe = AsyncMock(
        side_effect=ValueError("Dataset not found")
    )

    response_404 = client.get("/api/v1/datasets/nonexistent/download")
    assert response_404.status_code == 404
    assert "not found" in response_404.json()["detail"].lower()

    # Error during processing
    mock_dataset_handler.get_dataframe = AsyncMock(
        side_effect=Exception("Database connection failed")
    )

    response_500 = client.get(f"/api/v1/datasets/{dataset_id}/download")
    assert response_500.status_code == 500
    assert "error downloading dataset" in response_500.json()["detail"].lower()


def test_detect_and_decode_csv_with_bom() -> None:
    csv_with_bom = b"\xef\xbb\xbfName,Age\nJohn,25"
    result = detect_and_decode_csv(csv_with_bom, "test.csv")
    assert result == "Name,Age\nJohn,25"
    assert not result.startswith("\ufeff")


def test_detect_and_decode_csv_latin1() -> None:
    csv_latin1 = "Name,City\nJohn,São Paulo".encode("latin-1")
    result = detect_and_decode_csv(csv_latin1, "test.csv")
    assert "São Paulo" in result


def test_detect_and_decode_csv_empty() -> None:
    with pytest.raises(ValueError, match="empty"):
        detect_and_decode_csv(b"", "test.csv")


def test_detect_delimiter_semicolon() -> None:
    content = "Name;Age;City\nJohn;25;NYC"
    assert detect_delimiter(content) == ";"


def test_detect_delimiter_tab() -> None:
    content = "Name\tAge\tCity\nJohn\t25\tNYC"
    assert detect_delimiter(content) == "\t"


def test_detect_delimiter_defaults_to_comma() -> None:
    content = "Name\nJohn"
    assert detect_delimiter(content) == ","


def test_load_and_validate_csv_success() -> None:
    content = "Name,Age\nJohn,25\nJane,30"
    df = load_and_validate_csv(content, "test.csv")
    assert df.height == 2
    assert df.width == 2


def test_load_and_validate_csv_empty_data() -> None:
    content = "Name,Age"
    with pytest.raises(ValueError, match="no data rows"):
        load_and_validate_csv(content, "test.csv")


def test_load_and_validate_csv_windows_line_endings() -> None:
    content = "Name,Age\r\nJohn,25\r\nJane,30"
    df = load_and_validate_csv(content, "test.csv")
    assert df.height == 2
