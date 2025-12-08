# Copyright 2024 DataRobot, Inc.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
from __future__ import annotations

import asyncio
import base64
import io
import json
import os
import sys
import tempfile
import uuid
from copy import deepcopy
from pathlib import Path
from typing import Any, List, Union, cast

import chardet
import datarobot as dr
import pandas as pd
import polars as pl
import polars.dataframe.frame
from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    FastAPI,
    Form,
    HTTPException,
    Request,
    Response,
    UploadFile,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.openapi.utils import get_openapi
from fastapi.responses import StreamingResponse
from openai.types.chat.chat_completion_message_param import ChatCompletionMessageParam
from openai.types.chat.chat_completion_user_message_param import (
    ChatCompletionUserMessageParam,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from starlette.background import BackgroundTask

from utils.analyst_db import (
    AnalystDB,
    DatasetMetadata,
    DataSourceType,
    ExternalDataStoreNameDataSourceType,
    InternalDataSourceType,
    get_data_source_type,
)
from utils.api_exceptions import ExceptionBody
from utils.chat_dataset_helper import cleanup_message_datasets
from utils.data_analyst_telemetry import telemetry
from utils.data_connections.database.database_implementations import (
    get_external_database,
)
from utils.data_connections.database.database_interface import NoDatabaseOperator
from utils.data_connections.datarobot.datarobot_dataset_handler import (
    DatasetSparkRecipe,
    DataSourceRecipe,
)
from utils.datarobot_client import use_user_token
from utils.logging_helper import get_logger

sys.path.append(os.path.dirname(os.path.realpath(__file__)))

from utils.api import (
    AnalysisGenerationError,
    list_registry_datasets,
    load_registry_datasets,
    log_memory,
    process_data_and_update_state,
    register_remote_registry_datasets,
    run_complete_analysis,
    summarize_conversation,
    sync_data_sources_and_datasets,
)
from utils.constants import (
    ALTERNATIVE_LLM_BIG,
    CONTEXT_WARNING_THRESHOLD,
    MODEL_CONTEXT_WINDOW,
)
from utils.schema import (
    AnalystChatMessage,
    AnalystDataset,
    ChatCreate,
    ChatMessagePayload,
    ChatRequest,
    ChatResponse,
    ChatUpdate,
    DataDictionary,
    DataDictionaryResponse,
    DataRegistryDataset,
    DatasetCleansedResponse,
    DictionaryCellUpdate,
    EmptyResponse,
    ExternalDataSourcesSelection,
    ExternalDataStore,
    FileUploadResponse,
    GetBusinessAnalysisResult,
    LoadDatabaseRequest,
    RunAnalysisResult,
    RunChartsResult,
    RunDatabaseAnalysisResult,
    SupportedDataSourceTypes,
)
from utils.token_tracking import (
    TiktokenCountingStrategy,
    TokenUsageTracker,
    count_messages_tokens,
)

logger = get_logger()

MAX_EXCEL_ROWS = 50000  # Maximum rows to export to Excel to prevent memory issues


def detect_and_decode_csv(raw_bytes: bytes, filename: str) -> str:
    """
    Detect encoding and decode CSV content with BOM handling.
    Tries multiple strategies to ensure robust file handling.

    Raises ValueError with descriptive message on failure.
    """
    # Strip UTF-8 BOM if present
    if raw_bytes.startswith(b"\xef\xbb\xbf"):
        logger.info(f"Stripped UTF-8 BOM from '{filename}'")
        raw_bytes = raw_bytes[3:]

    if not raw_bytes:
        raise ValueError(f"File '{filename}' is empty or could not be read.")

    # Detect encoding
    detection_result = chardet.detect(raw_bytes)
    detected_encoding = detection_result.get("encoding")
    confidence = detection_result.get("confidence") or 0

    if detected_encoding:
        logger.info(
            f"Detected encoding for '{filename}': {detected_encoding} "
            f"(confidence: {confidence:.2f})"
        )

    # Try detected encoding first if confidence is reasonable
    if detected_encoding and confidence >= 0.7:
        try:
            return raw_bytes.decode(detected_encoding)
        except (UnicodeDecodeError, LookupError) as e:
            logger.warning(
                f"Failed to decode '{filename}' with detected encoding "
                f"{detected_encoding}: {e}"
            )

    # Try UTF-8 as fallback
    try:
        decoded = raw_bytes.decode("utf-8")
        if detected_encoding != "utf-8":
            logger.info(f"Successfully decoded '{filename}' using fallback: utf-8")
        return decoded
    except UnicodeDecodeError:
        pass

    raise ValueError(
        f"Unable to decode file '{filename}'. All encoding attempts failed."
    )


def detect_delimiter(content: str) -> str:
    """Detect CSV delimiter by checking first few lines. Defaults to comma on any ambiguity."""
    lines = content.split("\n")[:5]

    delimiters = {",": 0, ";": 0, "\t": 0, "|": 0}
    for line in lines:
        if not line.strip():
            continue
        for d in delimiters:
            delimiters[d] += line.count(d)

    best = max(delimiters, key=lambda k: delimiters[k])
    return best if delimiters[best] > 0 else ","


def load_and_validate_csv(decoded_content: str, filename: str) -> pl.DataFrame:
    """
    Parse CSV content and validate the resulting DataFrame.

    Raises ValueError with descriptive message on failure.
    """
    # Normalize line endings (old Mac CR → LF, Windows CRLF → LF)
    if "\r" in decoded_content:
        decoded_content = decoded_content.replace("\r\n", "\n").replace("\r", "\n")

    # Auto-detect delimiter
    separator = detect_delimiter(decoded_content)
    if separator != ",":
        logger.info(f"Detected delimiter '{separator}' for '{filename}'")

    try:
        df = pl.read_csv(
            io.StringIO(decoded_content),
            infer_schema_length=10000,
            low_memory=True,
            separator=separator,
        )
    except pl.exceptions.ComputeError as e:
        raise ValueError(f"Unable to parse CSV file '{filename}'. Error: {str(e)}")

    if df.height == 0:
        raise ValueError(f"CSV file '{filename}' contains only headers, no data rows.")

    # Check if ragged lines were truncated
    lines = [line for line in decoded_content.split("\n") if line.strip()]
    if len(lines) > 1:
        expected_rows = len(lines) - 1  # subtract header
        if df.height < expected_rows * 0.9:  # More than 10% loss
            logger.warning(
                f"CSV file '{filename}' had inconsistent row lengths. "
                f"Expected ~{expected_rows} rows, got {df.height}."
            )

    return df


async def get_database(user_id: str) -> AnalystDB:
    analyst_db = await AnalystDB.create(
        user_id=user_id,
        db_path=Path("/tmp"),
        dataset_db_name="datasets.db",
        chat_db_name="chat.db",
        data_source_db_name="datasources.db",
        user_recipe_db_name="recipe.db",
        use_persistent_storage=bool(os.environ.get("APPLICATION_ID")),
    )
    return analyst_db


# Dependency to provide the initialized database
async def get_initialized_db(request: Request) -> AnalystDB:
    if (
        not hasattr(request.state.session, "analyst_db")
        or request.state.session.analyst_db is None
    ):
        if not request.headers.get("x-user-email"):
            logger.error("x-user-email is required in order to initialize the database")
        raise HTTPException(
            status_code=400,
            detail="Database not initialized.",
        )

    return cast(AnalystDB, request.state.session.analyst_db)


# Initialize FastAPI app
app = FastAPI(
    title="Data Analyst API",
    description="""
    An intelligent API for data analysis that provides capabilities including:
    - Dataset management (upload CSV/Excel files, connect to databases, access the Data Registry)
    - Data cleansing and standardization
    - Data dictionary creation and management
    - Chat-based data analysis conversations
    - Python code generation
    - Chart creation
    - Business insights generation

    Available endpoint groups:
    - /api/v1/registry: Access Data Registry datasets
    - /api/v1/database: Database connection and table management
    - /api/v1/datasets: Upload, retrieve, and manage datasets
    - /api/v1/dictionaries: Manage data dictionaries
    - /api/v1/chats: Create and manage chat conversations for data analysis

    The API uses OpenAI's GPT models for intelligent analysis and response generation.
    """,
    version="1.0.0",
    contact={"name": "API Support", "email": "support@example.com"},
    license_info={
        "name": "Apache 2.0",
        "url": "https://www.apache.org/licenses/LICENSE-2.0.html",
    },
    debug=True,  # Stack traces will be exposed for 500 responses
)


script_name = os.environ.get("SCRIPT_NAME", "")
router = APIRouter(prefix=f"{script_name}/api/v1")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods
    allow_headers=["*"],  # Allows all headers
)


# Add custom OpenAPI schema
def custom_openapi() -> dict[str, Any]:
    if app.openapi_schema:
        return app.openapi_schema

    openapi_schema = get_openapi(
        title=app.title,
        version=app.version,
        description=app.description,
        routes=app.routes,
    )

    # Add security scheme
    openapi_schema["components"]["securitySchemes"] = {
        "ApiKeyAuth": {"type": "apiKey", "in": "header", "name": "X-API-Key"}
    }

    app.openapi_schema = openapi_schema
    return app.openapi_schema


app.openapi = custom_openapi  # type: ignore[method-assign]


class SessionState(object):
    _state: dict[str, Any]

    def __init__(self, state: dict[str, Any] | None = None):
        if state is None:
            state = {}
        super().__setattr__("_state", state)

    def __setattr__(self, key: Any, value: Any) -> None:
        self._state[key] = value

    def __getattr__(self, key: Any) -> Any:
        try:
            return self._state[key]
        except KeyError:
            message = "'{}' object has no attribute '{}'"
            raise AttributeError(message.format(self.__class__.__name__, key))

    def __delattr__(self, key: Any) -> None:
        del self._state[key]

    def update(self, state: dict[str, Any]) -> None:
        self._state.update(state)


session_store: dict[str, SessionState] = {}
session_lock = asyncio.Lock()


@app.middleware("http")
async def add_session_middleware(request: Request, call_next):  # type: ignore[no-untyped-def]
    request_methods = ["GET", "POST", "PUT", "PATCH", "DELETE"]
    if request.method in request_methods:
        # Initialize the session
        session_state, session_id, user_id = await _initialize_session(request)
        request.state.session = session_state

        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.replace("Bearer ", "")
            if request.state.session.datarobot_api_skoped_token != token:
                request.state.session.datarobot_api_skoped_token = token

        account_info = {}
        if not request.state.session.datarobot_account_info:
            try:
                if (
                    request.state.session.datarobot_api_token
                    or request.state.session.datarobot_api_skoped_token
                ):
                    with use_user_token(request):
                        reply = dr.client.get_client().get("account/info/")
                        account_info = reply.json()
            except Exception as e:
                logger.info(f"Error fetching account info: {e}")
        else:
            account_info = request.state.session.datarobot_account_info

        request.state.session.datarobot_account_info = account_info
        dr_uid = request.state.session.datarobot_account_info.get("uid")
        if session_id is None and dr_uid is not None:
            session_id = base64.b64encode(dr_uid.encode()).decode()
            user_id = dr_uid

        # Initialize database in the session
        if user_id:
            await _initialize_database(request, user_id)

    # Process the request
    response: Response = await call_next(request)

    if request.method in request_methods:
        # Set session cookie if needed
        if session_id:
            _set_session_cookie(
                response, user_id, session_id, request.cookies.get("session_fastapi")
            )

    return response


@telemetry.trace
@telemetry.meter
async def _initialize_session(
    request: Request,
) -> tuple[
    SessionState,
    str | None,
    str | None,
]:
    """Initialize the session state and return the session ID and user ID."""
    # Create a new session state with default values
    is_local_dev = os.environ.get("DEV_MODE", False)
    session_state = SessionState()
    empty_session_state: dict[str, Any] = {
        "datarobot_account_info": None,
        "datarobot_endpoint": os.environ.get("DATAROBOT_ENDPOINT"),
        "datarobot_api_token": os.environ.get("DATAROBOT_API_TOKEN")
        if is_local_dev
        else None,
        "datarobot_api_skoped_token": None,
        "analyst_db": None,
    }
    session_state.update(deepcopy(empty_session_state))

    # Try to get user ID from cookie
    user_id = None
    session_fastapi_cookie = request.cookies.get("session_fastapi")
    if session_fastapi_cookie:
        try:
            user_id = base64.b64decode(session_fastapi_cookie.encode()).decode()
        except Exception:
            pass  # If decoding fails, continue without user_id

    # Generate a new user ID if needed
    new_user_id = None
    email_header = request.headers.get("x-user-email")
    if email_header:
        new_user_id = str(uuid.uuid5(uuid.NAMESPACE_OID, email_header))[:36]

    # Determine session ID
    session_id = None
    if session_fastapi_cookie:
        session_id = session_fastapi_cookie
    elif new_user_id:
        session_id = base64.b64encode(new_user_id.encode()).decode()

    # Get or create session in store
    if session_id:
        async with session_lock:
            existing_session = session_store.get(session_id)
            if existing_session:
                return existing_session, session_id, user_id or new_user_id
            else:
                session_store[session_id] = session_state

    return session_state, session_id, user_id or new_user_id


async def _initialize_database(request: Request, user_id: str) -> None:
    """Initialize per-user database in the session if not already initialized."""
    if (
        not hasattr(request.state.session, "analyst_db")
        or request.state.session.analyst_db is None
    ):
        async with session_lock:
            request.state.session.analyst_db = await get_database(user_id)


def _set_session_cookie(
    response: Response,
    user_id: str | None,
    session_id: str,
    session_fastapi_cookie: str | None,
) -> None:
    """Set the session cookie if needed."""
    if user_id and not session_fastapi_cookie:
        encoded_uid = base64.b64encode(user_id.encode()).decode()
        response.set_cookie(key="session_fastapi", value=encoded_uid, httponly=True)
    elif not session_fastapi_cookie and not user_id:
        response.set_cookie(key="session_fastapi", value=session_id, httponly=True)


# Make this sync as the DR requests are synchronous, if async this would block.
@router.get("/registry/datasets")
def get_registry_datasets(
    request: Request, remote: bool = False, limit: int = 100
) -> list[DataRegistryDataset]:
    """Return all registry datasets

    Args:
        request (Request): _description_
        filter_downloadable (bool, optional): _description_. Defaults to False.
        limit (int, optional): _description_. Defaults to 100.

    Returns:
        list[DataRegistryDataset]: _description_
    """
    with use_user_token(request):
        return list_registry_datasets(remote=remote, limit=limit)


@router.get("/database/tables")
async def get_database_tables() -> list[str]:
    return await get_external_database().get_tables()


async def process_and_update(
    dataset_names: List[str],
    analyst_db: AnalystDB,
    datasource_type: DataSourceType,
) -> None:
    async for _ in process_data_and_update_state(
        dataset_names, analyst_db, datasource_type
    ):
        pass


@router.post("/datasets/upload")
async def upload_files(
    request: Request,
    background_tasks: BackgroundTasks,
    data_source: str | InternalDataSourceType | None = None,
    analyst_db: AnalystDB = Depends(get_initialized_db),
    files: List[UploadFile] | None = None,
    registry_ids: str | None = Form(None),
) -> list[FileUploadResponse]:
    normalized_data_source: InternalDataSourceType = (
        InternalDataSourceType.FILE
        if data_source is None
        else InternalDataSourceType(data_source)
        if isinstance(data_source, str)
        else data_source
    )
    dataset_names = []
    response: list[FileUploadResponse] = []
    if files:
        for file in files:
            try:
                file_size = file.size or 0
                contents = await file.read()
                if file.filename is None:
                    continue

                file_extension = os.path.splitext(file.filename)[1].lower()

                if file_extension == ".csv":
                    logger.info(f"Loading CSV: {file.filename}")
                    log_memory()

                    try:
                        decoded_content = detect_and_decode_csv(contents, file.filename)
                        df = load_and_validate_csv(decoded_content, file.filename)
                    except ValueError as e:
                        raise HTTPException(status_code=400, detail=str(e))
                    except Exception as e:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Failed to read CSV file '{file.filename}': {str(e)}",
                        )

                    log_memory()
                    dataset_name = os.path.splitext(file.filename)[0]
                    dataset = AnalystDataset(name=dataset_name, data=df)

                    # Register dataset with the database
                    await analyst_db.register_dataset(
                        dataset, InternalDataSourceType.FILE, file_size=file_size
                    )

                    # Add to processing queue
                    dataset_names.append(dataset.name)

                    file_response: FileUploadResponse = {
                        "filename": file.filename,
                        "content_type": file.content_type,
                        "size": len(contents),
                        "dataset_name": dataset_name,
                    }
                    response.append(file_response)

                elif file_extension in [".xlsx", ".xls"]:
                    base_name = os.path.splitext(file.filename)[0]

                    try:
                        excel_dataset = pl.read_excel(
                            io.BytesIO(contents), sheet_id=None
                        )
                    except Exception as e:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Unable to read Excel file '{file.filename}'. Error: {str(e)}",
                        )

                    if isinstance(excel_dataset, dict):
                        for sheet_name, data in excel_dataset.items():
                            dataset_name = f"{base_name}_{sheet_name}"
                            dataset = AnalystDataset(name=dataset_name, data=data)
                            await analyst_db.register_dataset(
                                dataset,
                                InternalDataSourceType.FILE,
                                file_size=file_size,
                            )
                            # Add to processing queue
                            dataset_names.append(dataset.name)

                            excel_sheet_response: FileUploadResponse = {
                                "filename": file.filename,
                                "content_type": file.content_type,
                                "size": len(contents),
                                "dataset_name": dataset_name,
                            }
                            response.append(excel_sheet_response)
                    elif isinstance(excel_dataset, pl.DataFrame):
                        dataset_name = base_name
                        dataset = AnalystDataset(name=dataset_name, data=excel_dataset)
                        await analyst_db.register_dataset(
                            dataset, InternalDataSourceType.FILE, file_size=file_size
                        )
                        # Add to processing queue
                        dataset_names.append(dataset.name)

                        excel_file_response: FileUploadResponse = {
                            "filename": file.filename,
                            "content_type": file.content_type,
                            "size": len(contents),
                            "dataset_name": dataset_name,
                        }
                        response.append(excel_file_response)
                else:
                    raise ValueError(f"Unsupported file type: {file_extension}")

            except Exception as e:
                error_response: FileUploadResponse = {
                    "filename": file.filename or "unknown_file",
                    "error": str(e),
                }
                response.append(error_response)

    # Process the data in the background (cleansing and dictionary generation)
    if dataset_names:
        background_tasks.add_task(
            process_and_update, dataset_names, analyst_db, InternalDataSourceType.FILE
        )

    if registry_ids:
        id_list: list[str] = json.loads(registry_ids)
        if id_list:
            with use_user_token(request):
                if normalized_data_source == InternalDataSourceType.REMOTE_REGISTRY:
                    dataframes, tasks = await register_remote_registry_datasets(
                        request, id_list, analyst_db
                    )
                else:
                    dataframes = await load_registry_datasets(id_list, analyst_db)
                    tasks = []
                for func, args, kwargs in tasks:
                    background_tasks.add_task(func, *args, **kwargs)
                dataset_names = [
                    dataset.name for dataset in dataframes if not dataset.error
                ]
                background_tasks.add_task(
                    process_and_update,
                    dataset_names,
                    analyst_db,
                    InternalDataSourceType.REMOTE_REGISTRY
                    if normalized_data_source == InternalDataSourceType.REMOTE_REGISTRY
                    else InternalDataSourceType.REGISTRY,
                )
                for dts in dataframes:
                    dts_response: FileUploadResponse = {
                        "dataset_name": dts.name,
                        "error": dts.error,
                    }
                    response.append(dts_response)

    return response


@router.post("/database/select")
async def load_from_database(
    data: LoadDatabaseRequest,
    background_tasks: BackgroundTasks,
    analyst_db: AnalystDB = Depends(get_initialized_db),
    sample_size: int = 1_000,
) -> list[str]:
    await asyncio.gather(
        *[
            analyst_db.register_dataset(
                AnalystDataset(name=table), InternalDataSourceType.DATABASE
            )
            for table in data.table_names
        ]
    )

    if data.table_names:
        background_tasks.add_task(
            get_and_process_tables, data.table_names, analyst_db, sample_size
        )

    return data.table_names


async def get_and_process_tables(
    table_names: list[str], analyst_db: AnalystDB, sample_size: int
) -> None:
    dataset_names = await get_external_database().get_data(
        *table_names, analyst_db=analyst_db, sample_size=sample_size
    )
    await process_and_update(dataset_names, analyst_db, InternalDataSourceType.DATABASE)


@router.get("/dictionaries")
async def get_dictionaries(
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> list[DataDictionaryResponse]:
    # Get datasets from the database
    dataset_names = await analyst_db.list_analyst_datasets()

    # Fetch dictionaries for each dataset
    dictionaries = []
    for name in dataset_names:
        dictionary = await analyst_db.get_data_dictionary(name)
        if dictionary and len(dictionary.column_descriptions):
            dictionaries.append(cast(DataDictionaryResponse, dictionary))
        else:
            dictionaries.append(
                DataDictionaryResponse(
                    name=name, column_descriptions=[], in_progress=True
                )
            )

    return dictionaries if dictionaries else []


@router.get("/datasets/{dataset_id}")
async def get_dataset_by_id(
    dataset_id: str,
    skip: int = 0,
    limit: int = 1000,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> DatasetCleansedResponse:
    """
    Get a dataset by its ID with pagination support.

    Args:
        dataset_id: The unique identifier of the dataset
        skip: Number of records to skip (for pagination)
        limit: Maximum number of records to return (for pagination)

    Returns:
        A response containing the dataset (similar to cleansed dataset format)

    Raises:
        HTTPException: If the dataset doesn't exist or cannot be retrieved
    """
    try:
        from utils.analyst_db import DatasetType

        df = await analyst_db.dataset_handler.get_dataframe(
            dataset_id,
            expected_type=DatasetType.ANALYST_RESULT_DATASET,
            max_rows=None,
        )

        if skip > 0 or limit > 0:
            df = df.slice(skip, limit)

        metadata = await analyst_db.dataset_handler.get_dataset_metadata(dataset_id)

        dataset = AnalystDataset(
            name=metadata.original_name,
            data=df,
        )

        # Return in the same format as cleansed dataset
        return DatasetCleansedResponse(
            dataset_name=metadata.original_name,
            cleaning_report=None,  # No cleaning report for analyst datasets
            dataset=dataset,
        )
    except ValueError as e:
        raise HTTPException(
            status_code=404, detail=f"Dataset '{dataset_id}' not found: {str(e)}"
        )
    except Exception as e:
        logger.error(f"Error retrieving dataset '{dataset_id}': {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Error retrieving dataset: {str(e)}"
        )


@router.get("/datasets/{dataset_id}/download")
async def download_dataset(
    dataset_id: str,
    analyst_db: AnalystDB = Depends(get_initialized_db),
    bom: bool = False,
) -> Response:
    """
    Download a dataset by ID as a CSV file.

    Args:
        dataset_id: The unique identifier of the dataset to download
        bom: Whether to include UTF-8 BOM for Excel compatibility (default: False)

    Returns:
        CSV file attachment

    Raises:
        HTTPException: If the dataset doesn't exist or cannot be retrieved
    """
    try:
        from utils.analyst_db import DatasetType

        df = await analyst_db.dataset_handler.get_dataframe(
            dataset_id,
            expected_type=DatasetType.ANALYST_RESULT_DATASET,
            max_rows=None,
        )

        csv_content = io.StringIO()
        df.write_csv(csv_content)

        csv_text = csv_content.getvalue()
        if bom:
            csv_text = "\ufeff" + csv_text

        response = Response(content=csv_text)
        response.headers["Content-Type"] = "text/csv; charset=utf-8"

        filename = f"analysis_result_dataset_{dataset_id[:8]}.csv"
        response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except ValueError as e:
        raise HTTPException(
            status_code=404, detail=f"Dataset '{dataset_id}' not found: {str(e)}"
        )
    except Exception as e:
        logger.error(f"Error downloading dataset '{dataset_id}': {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Error downloading dataset: {str(e)}"
        )


@router.get("/datasets/{name}/metadata")
async def get_dataset_metadata(
    name: str, analyst_db: AnalystDB = Depends(get_initialized_db)
) -> DatasetMetadata:
    """
    Get metadata for a dataset by name from the database.

    Args:
        name: The name of the dataset

    Returns:
        A dictionary containing dataset metadata including type, creation date, columns, and row count

    Raises:
        HTTPException: If the dataset doesn't exist or cannot be retrieved
    """
    try:
        return await analyst_db.get_dataset_metadata(name)
    except ValueError as e:
        raise HTTPException(
            status_code=404, detail=f"Dataset metadata not found: {str(e)}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error retrieving dataset metadata: {str(e)}"
        )


@router.get("/datasets/{name}/cleansed")
async def get_cleansed_dataset(
    name: str,
    skip: int = 0,
    limit: int = 10000,
    search: str | None = None,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> DatasetCleansedResponse:
    """
    Get a cleansed dataset by name from the database with pagination and search support.

    Args:
        name: The name of the dataset
        skip: Number of records to skip (for pagination)
        limit: Maximum number of records to return (for pagination)
        search: Optional search term to filter columns by name (raw data view)

    Returns:
        A dictionary containing the cleaning report (if available) and the dataset as a list of records.

    Raises:
        HTTPException: If the dataset doesn't exist or cannot be retrieved
    """
    try:
        ds_display = await analyst_db.get_dataset(name)

        # Initialize the response structure
        response = DatasetCleansedResponse(
            dataset_name=name,
            cleaning_report=None,
            dataset=None,
        )

        # Attempt to fetch the cleaning report
        try:
            ds_display_cleansed = await analyst_db.get_cleansed_dataset(name)
            response.cleaning_report = ds_display_cleansed.generate_cleaning_report()

        except ValueError:
            # Cleaning report is not available
            response.cleaning_report = None

        # Convert the dataset to a DataFrame
        df_display = ds_display.to_df()

        # Apply search filter if provided - filter columns by name
        if search and search.strip():
            search_term = search.strip().lower()
            original_columns = list(df_display.columns)
            matching_columns = [
                col for col in original_columns if search_term in col.lower()
            ]
            if matching_columns:
                # Select only the matching columns
                try:
                    df_display = df_display.select(matching_columns)
                except Exception as e:
                    logger.error(f"Error selecting columns: {e}")
                    # Fallback to original dataframe if selection fails
                    pass
            else:
                # No matching columns, create empty dataframe
                df_display = df_display.select([]).limit(0)

        # Apply pagination (skip and limit)
        if skip > 0 or limit > 0:
            df_display = df_display.slice(skip, limit)

        # Create an instance of AnalystDataset
        dataset = AnalystDataset(
            name=name,
            columns=df_display.columns,
            data=df_display.to_dicts(),  # Convert rows to a list of dictionaries
        )

        # Add the dataset to the response
        response.dataset = dataset

        return response

    except ValueError as e:
        raise HTTPException(
            status_code=404, detail=f"Dataset '{name}' not found: {str(e)}"
        )
    except Exception as e:
        logger.error(f"Error retrieving cleansed dataset '{name}': {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Error retrieving cleansed dataset: {str(e)}"
        )


@router.delete("/datasets", status_code=200)
async def delete_datasets(
    request: Request,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> None:
    await analyst_db.delete_all_tables()


@router.delete("/dictionaries/{name}", status_code=200)
async def delete_dictionary(
    name: str, analyst_db: AnalystDB = Depends(get_initialized_db)
) -> None:
    await analyst_db.delete_table(name)


@router.get("/dictionaries/{name}/download")
async def download_dictionary(
    name: str,
    analyst_db: AnalystDB = Depends(get_initialized_db),
    bom: bool = False,
) -> Response:
    """
    Download a dictionary as a CSV file.

    Args:
        name: Name of the dataset whose dictionary to download

    Returns:
        CSV file attachment
    """
    dictionary = await analyst_db.get_data_dictionary(name)

    if not dictionary:
        raise HTTPException(status_code=404, detail=f"Dictionary '{name}' not found")

    # Convert the dictionary to a DataFrame
    df = dictionary.to_application_df()

    # Convert to CSV
    csv_content = io.StringIO()
    df.write_csv(csv_content)

    # Create response with CSV attachment
    csv_text = csv_content.getvalue()
    if bom:
        csv_text = (
            "\ufeff" + csv_text
        )  # Prefix UTF-8 BOM for Excel compatibility with international characters
    response = Response(content=csv_text)
    response.headers["Content-Type"] = "text/csv; charset=utf-8"

    return response


@router.patch("/dictionaries/{name}/cells")
async def update_dictionary_cell(
    name: str,
    update: DictionaryCellUpdate,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> DataDictionary:
    dictionary = await analyst_db.get_data_dictionary(name)

    if not dictionary:
        raise HTTPException(status_code=404, detail=f"Dictionary '{name}' not found")

    if update.rowIndex < 0 or update.rowIndex >= len(dictionary.column_descriptions):
        raise HTTPException(
            status_code=400, detail=f"Row index {update.rowIndex} is out of range"
        )

    column_description = dictionary.column_descriptions[update.rowIndex]
    if hasattr(column_description, update.field):
        setattr(column_description, update.field, update.value)
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Field '{update.field}' not found in dictionary row",
        )

    await analyst_db.register_data_dictionary(dictionary, clobber=True)

    return dictionary


@router.post("/chats")
async def create_chat(
    chat: ChatCreate,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> dict[str, str]:
    """Create a new chat with optional data source"""

    chat_id = await analyst_db.create_chat(
        chat_name=chat.name,
        data_source=chat.data_source,
    )

    return {"id": chat_id}


@router.get("/chats")
async def get_chats(
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> list[dict[str, Any]]:
    """Get all chats"""
    chat_list = await analyst_db.get_chat_list()

    return [
        {
            "id": chat["id"],
            "name": chat["name"],
            "data_source": chat.get("data_source", "catalog"),
            "created_at": chat["created_at"],
        }
        for chat in chat_list
    ]


@router.get("/chats/{chat_id}")
async def get_chat(
    chat_id: str, analyst_db: AnalystDB = Depends(get_initialized_db)
) -> ChatResponse:
    """Get a specific chat by ID"""
    chat = await analyst_db.get_chat_messages(chat_id=chat_id)

    return {
        "id": chat_id,
        "messages": chat,
    }


@router.put("/chats/{chat_id}")
async def update_chat(
    chat_id: str, chat: ChatUpdate, analyst_db: AnalystDB = Depends(get_initialized_db)
) -> dict[str, str]:
    """Update a chat's name and/or data source"""
    response_messages = []

    # Update chat name if provided
    if chat.name:
        await analyst_db.rename_chat(chat_id, chat.name)
        response_messages.append("renamed")

    # Update data source if provided
    if chat.data_source:
        await analyst_db.update_chat_data_source(chat_id, chat.data_source)
        response_messages.append("updated data source")

    if not response_messages:
        return {"message": f"No changes made to chat with ID {chat_id}"}

    return {
        "message": f"Chat with ID {chat_id} was {' and '.join(response_messages)} successfully"
    }


@router.delete("/chats/{chat_id}", status_code=200)
async def delete_chat(
    chat_id: str, analyst_db: AnalystDB = Depends(get_initialized_db)
) -> dict[str, str]:
    """Delete a chat"""
    # Get all messages to clean up datasets
    messages = await analyst_db.get_chat_messages(chat_id=chat_id)

    # Clean up datasets for all messages
    for message in messages:
        await cleanup_message_datasets(analyst_db, message)

    # Delete the chat
    await analyst_db.delete_chat(chat_id=chat_id)

    return {"message": f"Chat with ID {chat_id} deleted successfully"}


@router.get("/chats/{chat_id}/messages")
async def get_chat_messages(
    chat_id: str,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> list[AnalystChatMessage]:
    """Get messages for a specific chat"""

    chat = await analyst_db.get_chat_messages(chat_id=chat_id)

    return chat


@router.delete("/chats/messages/{message_id}")
async def delete_chat_message(
    request: Request,
    message_id: str,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> list[AnalystChatMessage]:
    """Delete a specific message"""
    try:
        message = await analyst_db.get_chat_message(message_id=message_id)
        if not message:
            raise HTTPException(
                status_code=404, detail=f"Message with ID {message_id} not found"
            )
        else:
            # Clean up associated datasets before deleting the message
            await cleanup_message_datasets(analyst_db, message)

            await analyst_db.delete_chat_message(message_id=message_id)
            messages = await analyst_db.get_chat_messages(
                chat_id=message.chat_id,
            )
            return cast(list[AnalystChatMessage], list(messages))
    except Exception as e:
        logger.error(f"Error deleting message: {str(e)}")

        return cast(list[AnalystChatMessage], [])


@router.get("/chats/{chat_id}/messages/{message_id}")
async def get_chat_message(
    chat_id: str,
    message_id: str,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> AnalystChatMessage:
    """Get a specific message by ID from a specific chat"""
    try:
        message = await analyst_db.get_chat_message(message_id=message_id)
        if not message:
            raise HTTPException(
                status_code=404, detail=f"Message with ID {message_id} not found"
            )

        # Verify the message belongs to the specified chat
        if message.chat_id != chat_id:
            raise HTTPException(
                status_code=404,
                detail=f"Message with ID {message_id} not found in chat {chat_id}",
            )

        return message
    except Exception as e:
        logger.error(f"Error getting message: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Error retrieving message: {str(e)}"
        )


@router.post("/chats/messages")
async def create_new_chat_message(
    request: Request,
    payload: ChatMessagePayload,
    background_tasks: BackgroundTasks,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> dict[str, Union[str, list[AnalystChatMessage], None]]:
    """Create a new chat and post a message to it"""

    # Create a new chat
    chat_id = await analyst_db.create_chat(
        chat_name=payload.chatName,
        data_source=payload.data_source,
    )

    # Create the user message
    user_message = AnalystChatMessage(
        role="user", content=payload.message, components=[]
    )

    message_id = await analyst_db.add_chat_message(
        chat_id=chat_id,
        message=user_message,
    )

    # Create valid messages for the chat request
    valid_messages: list[ChatCompletionMessageParam] = [
        user_message.to_openai_message_param()
    ]

    # Add the current message
    valid_messages.append(
        ChatCompletionUserMessageParam(role="user", content=payload.message)
    )

    # Create the chat request
    chat_request = ChatRequest(messages=valid_messages)

    # Run the analysis in the background
    # Running the sync version whill get run in a different thread as the async version,
    # despite the name, blocks for significant period. (When that gets fixed, we can change this.)
    background_tasks.add_task(
        run_complete_analysis_task,
        chat_request,
        payload.data_source,
        analyst_db,
        chat_id,
        message_id,
        payload.enable_chart_generation,
        payload.enable_business_insights,
        request,
    )

    chat_list = await analyst_db.get_chat_list()
    chat_name = next((n["name"] for n in chat_list if n["id"] == chat_id), None)
    messages = await analyst_db.get_chat_messages(chat_id=chat_id)

    chat = {
        "id": chat_id,
        "name": chat_name,
        "messages": messages,
    }

    return chat


@router.post("/chats/{chat_id}/messages")
async def create_chat_message(
    request: Request,
    chat_id: str,
    payload: ChatMessagePayload,
    background_tasks: BackgroundTasks,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> dict[str, Union[str, list[AnalystChatMessage], None]]:
    """Post a message to a specific chat"""
    messages = await analyst_db.get_chat_messages(chat_id=chat_id)

    # Check if any message is in progress
    in_progress = any(message.in_progress for message in messages)

    # Check if cancelled
    if not in_progress:
        # Check if there's an existing summary (find last system message)
        last_summary_idx = None
        for i in range(len(messages) - 1, -1, -1):  # Iterate backwards through indices
            if messages[i].role == "system":
                last_summary_idx = i
                break

        # Build context: if summary exists, use [summary] + messages_after, else all messages
        if last_summary_idx is not None:
            context_messages = messages[last_summary_idx:]
            logger.info(
                f"[chat_id={chat_id}] Using summary + {len(context_messages) - 1} messages after"
            )
        else:
            context_messages = messages

        # Create valid messages for the chat request
        valid_messages: list[ChatCompletionMessageParam] = [
            msg.to_openai_message_param()
            for msg in context_messages
            if msg.content.strip()
        ]

        # Add the current message
        valid_messages.append(
            ChatCompletionUserMessageParam(role="user", content=payload.message)
        )

        # Check context usage
        tokens_used = count_messages_tokens(valid_messages, ALTERNATIVE_LLM_BIG)
        usage_pct = (tokens_used / MODEL_CONTEXT_WINDOW) * 100

        logger.info(
            f"[chat_id={chat_id}] Context: {tokens_used:,}/{MODEL_CONTEXT_WINDOW:,} tokens ({usage_pct:.1f}%)"
        )

        # Create and store the user message first (before summarization)
        user_message = AnalystChatMessage(
            role="user", content=payload.message, components=[]
        )
        message_id = await analyst_db.add_chat_message(
            chat_id=chat_id, message=user_message
        )
        user_message.id = message_id

        # Trigger summarization if over threshold
        if tokens_used >= CONTEXT_WARNING_THRESHOLD:
            logger.warning(
                f"[chat_id={chat_id}] ⚠️  Context at {usage_pct:.1f}% - creating summary"
            )

            # Create new system message with in_progress=True
            summary_message = AnalystChatMessage(
                role="system",
                content="Summarizing conversation...",
                components=[],
                in_progress=True,
            )
            summary_message.id = await analyst_db.add_chat_message(
                chat_id=chat_id, message=summary_message
            )

            # Determine what to summarize: from last system message (if exists) to now, or all messages
            if last_summary_idx is not None:
                # Summarize from last system message onwards
                messages_to_summarize = [
                    msg.to_openai_message_param()
                    for msg in messages[last_summary_idx:]
                    if msg.content.strip()
                ]
            else:
                # Summarize all messages
                messages_to_summarize = [
                    msg.to_openai_message_param()
                    for msg in messages
                    if msg.content.strip()
                ]

            # Create token tracker for summarization
            summarization_tracker = TokenUsageTracker(
                strategy=TiktokenCountingStrategy()
            )

            try:
                summary_text = await summarize_conversation(
                    messages_to_summarize, token_tracker=summarization_tracker
                )

                logger.info(
                    f"[chat_id={chat_id}] Summarization token usage: "
                    f"{summarization_tracker.prompt_tokens} prompt + "
                    f"{summarization_tracker.completion_tokens} completion = "
                    f"{summarization_tracker.total_tokens} total tokens"
                )

                # Update the summary message with actual content
                summary_message.content = summary_text
                summary_message.in_progress = False

                await analyst_db.update_chat_message(
                    message_id=summary_message.id,
                    message=summary_message,
                )

                logger.info(
                    f"[chat_id={chat_id}] Summary stored ({len(summary_text)} chars)"
                )

                # Rebuild context using the new summary
                summary_param = summary_message.to_openai_message_param()
                valid_messages = [
                    summary_param,
                    valid_messages[-1],
                ]  # summary + current message

            except Exception as e:
                logger.error(f"[chat_id={chat_id}] Failed to create summary: {e}")
                # Mark summary as failed
                summary_message.content = "Failed to create summary"
                summary_message.in_progress = False
                summary_message.error = str(e)
                await analyst_db.update_chat_message(
                    message_id=summary_message.id,
                    message=summary_message,
                )
                # Continue without summary

        # Create the chat request
        chat_request = ChatRequest(messages=valid_messages)

        # Run the analysis in the background
        # Running the sync version whill get run in a different thread as the async version,
        # despite the name, blocks for significant period. (When that gets fixed, we can change this.)
        background_tasks.add_task(
            run_complete_analysis_task,
            chat_request,
            payload.data_source,
            analyst_db,
            chat_id,
            message_id,
            payload.enable_chart_generation,
            payload.enable_business_insights,
            request,
        )

    chat_list = await analyst_db.get_chat_list()
    chat_name = next((n["name"] for n in chat_list if n["id"] == chat_id), None)
    messages = await analyst_db.get_chat_messages(chat_id=chat_id)

    chat = {
        "id": chat_id,
        "name": chat_name,
        "messages": messages,
    }

    return chat


@router.get("/chats/{chat_id}/messages/download/")
async def save_chat_messages(
    request: Request,
    chat_id: str,
    analyst_db: AnalystDB = Depends(get_initialized_db),
    message_id: str | None = None,
) -> StreamingResponse:
    """
    This API controller saves a chat ID to an excel spreadsheet which
    saves key information, then is streamed back to the user.
    """
    temp_files: list[str] = []

    chat_messages: List[AnalystChatMessage] = await analyst_db.get_chat_messages(
        chat_id=chat_id
    )

    # If a specific message_id is provided, filter messages to include
    # only that user message and its following assistant response.
    if message_id:
        idx = next((i for i, m in enumerate(chat_messages) if m.id == message_id), None)
        if idx is None or chat_messages[idx].role != "user":
            raise HTTPException(detail="User message not found", status_code=404)

        # Find the following assistant message
        assistant_message = None
        for i in range(idx + 1, len(chat_messages)):
            if chat_messages[i].role == "assistant":
                assistant_message = chat_messages[i]
                break

        # Include user message and assistant response if found
        if assistant_message:
            filtered_messages = [chat_messages[idx], assistant_message]
        else:
            filtered_messages = [chat_messages[idx]]

        chat_messages = filtered_messages
    if not chat_messages:
        # Create an empty workbook with basic structure for empty chats
        analysis_workbook = Workbook()
        # Remove the default sheet
        analysis_workbook.remove(analysis_workbook.active)

        # Create a single sheet indicating the chat is empty
        empty_sheet = analysis_workbook.create_sheet("Empty Chat")
        empty_sheet["A1"] = "Chat Export"
        empty_sheet["A3"] = "This chat contains no messages to export yet."

        output = io.BytesIO()
        analysis_workbook.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if any(msg.in_progress for msg in chat_messages):
        raise HTTPException(
            detail="Cannot download while a chat is in progress.", status_code=425
        )
    analysis_workbook = Workbook()
    data_sheets_count = 0
    charts_sheets_count = 0
    report_sheets_count = 0

    # Remove the initial default sheet created by Workbook()
    analysis_workbook.remove(analysis_workbook.active)

    for i, chat_message in enumerate(chat_messages):
        # Skip system messages (summarization messages)
        if chat_message.role == "system":
            continue

        if chat_message.role == "assistant":
            # Handle Analysis Report sheet
            report_sheets_count += 1
            report_sheet_name = (
                "Sheet" if report_sheets_count == 1 else f"Sheet {report_sheets_count}"
            )
            report_sheet = analysis_workbook.create_sheet(report_sheet_name)

            report_sheet["A1"] = "Analysis Report"

            # Find the previous user message by searching backwards
            user_question = "No question found"
            for j in range(i - 1, -1, -1):
                if chat_messages[j].role == "user":
                    user_question = chat_messages[j].content
                    break

            report_sheet["A3"] = "Question"
            report_sheet["A4"] = user_question
            report_sheet["A6"] = "Answer"
            report_sheet["A7"] = chat_message.content

            business_components: list[GetBusinessAnalysisResult] = [
                component
                for component in chat_message.components
                if isinstance(component, GetBusinessAnalysisResult)
            ]
            for index, business_component in enumerate(business_components):
                cell_index = index + 1  # Excel uses 1 indexing
                report_sheet.cell(9, cell_index).value = "Bottom Line"
                report_sheet.cell(10, cell_index).value = business_component.bottom_line

                report_sheet.cell(12, cell_index).value = "Additional Insights"
                report_sheet.cell(
                    13, cell_index
                ).value = business_component.additional_insights

                report_sheet.cell(15, cell_index).value = "Follow-up Questions:"
                for q_index, followup_question in enumerate(
                    business_component.follow_up_questions
                ):
                    report_sheet.cell(
                        16 + q_index, cell_index
                    ).value = followup_question

            # Handle Data sheets
            run_analysis_components: List[
                RunAnalysisResult | RunDatabaseAnalysisResult
            ] = [
                component
                for component in chat_message.components
                if isinstance(component, (RunAnalysisResult, RunDatabaseAnalysisResult))
            ]
            for run_analysis_component in run_analysis_components:
                # Load dataset from storage (dataset field is excluded from serialization)
                if not run_analysis_component.dataset_id:
                    continue

                try:
                    from utils.analyst_db import DatasetType

                    df = await analyst_db.dataset_handler.get_dataframe(
                        run_analysis_component.dataset_id,
                        expected_type=DatasetType.ANALYST_RESULT_DATASET,
                        max_rows=None,
                    )
                    metadata = await analyst_db.dataset_handler.get_dataset_metadata(
                        run_analysis_component.dataset_id
                    )
                    dataset_to_export = AnalystDataset(
                        name=metadata.original_name,
                        data=df,
                    )
                except Exception as e:
                    logger.error(
                        f"Failed to load dataset {run_analysis_component.dataset_id}: {e}"
                    )
                    continue

                data_sheets_count += 1
                data_sheet_name = (
                    "Data" if data_sheets_count == 1 else f"Data {data_sheets_count}"
                )
                data_sheet = analysis_workbook.create_sheet(data_sheet_name)

                try:
                    dataset: polars.dataframe.frame.DataFrame = (
                        dataset_to_export.data.df
                    )

                    # Convert to pandas with error handling for large datasets
                    pandas_df = dataset.to_pandas()

                    # Add size check to prevent memory issues and Excel limits
                    original_rows = pandas_df.shape[0]
                    if original_rows > MAX_EXCEL_ROWS:
                        logger.warning(
                            f"Dataset too large ({original_rows} rows), truncating to {MAX_EXCEL_ROWS} rows"
                        )
                        pandas_df = pandas_df.head(MAX_EXCEL_ROWS)
                        # Add a notice row at the top of the sheet
                        data_sheet.append(
                            [
                                f"NOTICE: Dataset truncated from {original_rows:,} to {MAX_EXCEL_ROWS:,} rows due to Excel limitations"
                            ]
                        )
                        data_sheet.append([])

                    for r in dataframe_to_rows(pandas_df, index=False, header=True):
                        data_sheet.append(r)

                except Exception as e:
                    logger.error(f"Failed to process dataset: {e}")
                    # Create error sheet instead of crashing
                    data_sheet["A1"] = "Dataset Processing Error"
                    data_sheet["A2"] = f"Error: {str(e)}"

            # Handle Charts sheets
            run_charts_components: List[RunChartsResult] = [
                component
                for component in chat_message.components
                if isinstance(component, RunChartsResult)
            ]
            for run_chart_component in run_charts_components:
                for f, js in [
                    (run_chart_component.fig1, run_chart_component.fig1_json),
                    (run_chart_component.fig2, run_chart_component.fig2_json),
                ]:
                    if not js or not f:
                        continue
                    charts_sheets_count += 1
                    charts_sheet_name = f"Chart {charts_sheets_count}"
                    charts_sheet = analysis_workbook.create_sheet(charts_sheet_name)

                    # Save the chart data by removing the some keys from this field
                    # We don't know what exactly is going to be graphed, so this is an easy way of covering
                    # ourselves
                    try:
                        parsed_json = json.loads(js)
                        data_list = parsed_json.get("data", [])
                        if data_list and len(data_list) > 0:
                            fig_json = data_list[
                                0
                            ].copy()  # Create copy to avoid modifying original
                            [fig_json.pop(k, None) for k in ["marker", "name", "type"]]
                            chart_df = pd.DataFrame(fig_json)
                            for r in dataframe_to_rows(
                                chart_df, index=False, header=True
                            ):
                                charts_sheet.append(r)

                        with tempfile.NamedTemporaryFile(
                            suffix=".png", delete=False
                        ) as tmpfile:
                            f.write_image(tmpfile.name)
                            img = XLImage(tmpfile.name)
                            charts_sheet.add_image(img, "F3")
                            temp_files.append(tmpfile.name)
                    except (
                        json.JSONDecodeError,
                        KeyError,
                        IndexError,
                        ValueError,
                    ) as e:
                        logger.warning(f"Failed to process chart data: {e}")
                        # Create error sheet instead of crashing
                        charts_sheet["A1"] = "Chart Processing Error"
                        charts_sheet["A2"] = f"Error: {str(e)}"
                    except Exception as e:
                        logger.error(f"Unexpected error processing chart: {e}")
                        continue  # Skip this chart but continue processing
    output = io.BytesIO()
    analysis_workbook.save(output)
    output.seek(0)

    # Create background task to cleanup temporary files
    def cleanup_files(file_paths: List[str]) -> None:
        for fp in file_paths:
            if os.path.exists(fp):
                os.remove(fp)

    background_task = BackgroundTask(cleanup_files, temp_files) if temp_files else None

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=background_task,
    )


async def run_complete_analysis_task(
    chat_request: ChatRequest,
    data_source: str,
    analyst_db: AnalystDB,
    chat_id: str,
    message_id: str,
    enable_chart_generation: bool,
    enable_business_insights: bool,
    request: Request,
) -> None:
    """Run the complete analysis pipeline"""
    source = get_data_source_type(data_source)
    logger.debug(
        "Running analysis for user.",
        extra={
            "data_source": data_source,
            "user_id": analyst_db.user_id,
        },
    )
    dataset_metadata = []
    if source in [InternalDataSourceType.REGISTRY, InternalDataSourceType.FILE]:
        dataset_metadata = (
            await analyst_db.list_analyst_dataset_metadata(
                InternalDataSourceType.REGISTRY
            )
        ) + (
            await analyst_db.list_analyst_dataset_metadata(InternalDataSourceType.FILE)
        )
    else:
        dataset_metadata = await analyst_db.list_analyst_dataset_metadata(source)

    run_analysis_iterator = run_complete_analysis(
        chat_request=chat_request,
        data_source=source,
        dataset_metadata=dataset_metadata,
        analyst_db=analyst_db,
        chat_id=chat_id,
        message_id=message_id,
        enable_chart_generation=enable_chart_generation,
        enable_business_insights=enable_business_insights,
        request=request,
    )

    async for message in run_analysis_iterator:
        if isinstance(message, AnalysisGenerationError):
            break
        else:
            pass


# Sync as this is a long blocking request
@router.get("/available-external-data-stores")
def get_available_external_data_stores(
    request: Request, analyst_db: AnalystDB = Depends(get_initialized_db)
) -> list[ExternalDataStore]:
    """List all available datastores. (An available datastore
    (a) has datasources configured and (b) has a supported driver.)

    Args:
        request (Request): HTTP request.
        limit (int): Maximum entries to retrieve.
        offset (int): Number of entries to skip.

    Returns:
        list[ExternalDataStore]: The given page of datastores.
    """
    with use_user_token(request):
        return asyncio.run(
            DataSourceRecipe.list_available_datastores(analyst_db.user_id)
        )


@router.put(
    "/external-data-stores/{external_data_store_id}/external-data-sources/",
    responses={404: {"model": ExceptionBody}},
)
async def register_external_data_sources(
    request: Request,
    selected_datasource_ids: ExternalDataSourcesSelection,
    external_data_store_id: str,
    background_tasks: BackgroundTasks,
    analyst_db: AnalystDB = Depends(get_initialized_db),
) -> EmptyResponse:
    """
    Add data sources for a datastore, registering them as datasets in the app.

    Args:
        request (Request): The request
        selected_datasource_ids (ExternalDataSourcesSelection): Select data sources for a datasource
        external_data_store_id (str): Select the data store
        background_tasks (BackgroundTasks): Background tasks
        analyst_db (AnalystDB, optional): The database. Defaults to Depends(get_initialized_db).

    Returns:
        ExternalDataStore: The updated data store
    """
    logger.debug(
        "PUT /external-data-stores/%s/external-data-sources/", external_data_store_id
    )
    with use_user_token(request):
        name = await DataSourceRecipe.get_canonical_name_for_datastore_id(
            external_data_store_id
        )
        if not name:
            raise HTTPException(
                404, f"DataStore {external_data_store_id} does not exist."
            )

        background_tasks.add_task(
            update_data_sources_for_data_store,
            request,
            selected_datasource_ids,
            external_data_store_id,
            analyst_db,
        )

        logger.debug(
            "Registering data sources for data store %s (%s).",
            name,
            external_data_store_id,
        )
        _, tasks = await sync_data_sources_and_datasets(
            request=request,
            canonical_name=name,
            analyst_db=analyst_db,
            data_store_id=external_data_store_id,
            selected_datasource_ids=selected_datasource_ids,
        )

        logger.debug(
            "Adding background tasks for data store %s.", external_data_store_id
        )
        for f, arg, kwargs in tasks:
            background_tasks.add_task(f, *arg, **kwargs)
        background_tasks.add_task(
            process_and_update,
            [d.path for d in selected_datasource_ids.selected_data_sources],
            analyst_db,
            ExternalDataStoreNameDataSourceType.from_name(name),
        )
    return EmptyResponse()


async def update_data_sources_for_data_store(
    request: Request,
    selected_datasource_ids: ExternalDataSourcesSelection,
    external_data_store_id: str,
    analyst_db: AnalystDB,
) -> None:
    with use_user_token(request):
        logger.debug("Loading recipe for datasources %s.")
        recipe = await DataSourceRecipe.load_or_create(
            analyst_db, external_data_store_id
        )
        logger.debug("Updating recipe with data sources for %s", external_data_store_id)
        await recipe.select_data_sources(selected_datasource_ids.selected_data_sources)


@router.get("/supported-data-source-types")
async def get_supported_datasource_types(request: Request) -> SupportedDataSourceTypes:
    types: list[InternalDataSourceType] = [
        InternalDataSourceType.FILE,
        InternalDataSourceType.REGISTRY,
    ]
    if not isinstance(get_external_database(), NoDatabaseOperator):
        types.append(InternalDataSourceType.DATABASE)
    if DatasetSparkRecipe.should_use_spark_recipe():
        types.append(InternalDataSourceType.REMOTE_REGISTRY)
    return SupportedDataSourceTypes(supported_types=[t.value for t in types])


@router.get("/user/datarobot-account")
async def get_datarobot_account(
    request: Request,
) -> dict[str, Any]:
    api_token = request.state.session.datarobot_api_token
    skoped_token = request.state.session.datarobot_api_skoped_token

    truncated_api_token = None
    if api_token:
        truncated_api_token = f"****{api_token[-4:]}"

    truncated_skoped_token = None
    if skoped_token:
        truncated_skoped_token = f"****{skoped_token[-4:]}"

    return {
        "datarobot_account_info": request.state.session.datarobot_account_info,
        "datarobot_api_token": truncated_api_token,
        "datarobot_api_skoped_token": truncated_skoped_token,
    }


@router.post("/user/datarobot-account")
async def store_datarobot_account(
    request: Request,
) -> dict[str, Any]:
    request_data = await request.json()

    if "api_token" in request_data and request_data["api_token"]:
        request.state.session.datarobot_api_token = request_data["api_token"]

    return {"success": True}


app.include_router(router)
